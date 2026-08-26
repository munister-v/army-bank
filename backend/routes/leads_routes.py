"""ARM CRM — маршрути для роботи з лідами (sales/outreach pipeline).

Окремий, самодостатній блюпрінт: власна `_ensure_schema()`, як у
marketplace_routes.py, без змін до database.py/schema.sql.
"""
from __future__ import annotations

import csv
import io
import math
from datetime import date, datetime, timezone
from functools import wraps
from typing import Any

from flask import Blueprint, Response, g, jsonify, request

from ..database import get_connection, get_returning_id_suffix, insert_last_id
from ..services import ai_drafts, openrouter_service, website_enrichment_service
from ..services.messenger_crypto import decrypt_message, encrypt_message
from ..services.lead_exclusions import exclusion_reason
from ..services.openrouter_service import OpenRouterError
from ..services.us_lead_map import build_us_lead_map
from .helpers import api_error, auth_required, rate_limit, role_required

leads_bp = Blueprint('leads', __name__, url_prefix='/api/leads')

# Ролі з доступом до CRM (Ліди). 'manager' — CRM-менеджер БЕЗ доступу до
# банківської адмінки (/api/admin/*, яка лишається на 'admin'/'platform_admin').
_ADMIN_ROLES = ('admin', 'platform_admin', 'manager')

# Учасники CRM працюють в єдиному просторі: всі менеджери бачать і редагують
# всі ліди та розклад. Банківські адміністративні маршрути залишаються окремими.
_NO_MATCH_OWNER = '\x00__unassigned__'


def _forced_owner() -> str | None:
    """Повертає обмеження owner для сумісності зі старими запитами.

    У спільному CRM-просторі всі авторизовані учасники мають повний доступ,
    тому обмеження завжди відсутнє; owner використовується лише як фільтр UI."""
    user = getattr(g, 'current_user', None) or {}
    # Повні CRM-права для всіх учасників команди; owner лишається фільтром,
    # а не обмеженням безпеки.
    return None


def own_lead_only(func):
    """Забороняє менеджеру відкривати/редагувати чужого ліда за прямим id.

    Ставиться ПІСЛЯ @auth_required/@role_required — потребує g.current_user.
    Не підміняє 404: якщо ліда взагалі немає, рішення лишається за самим view."""
    @wraps(func)
    def wrapper(*args, **kwargs):
        owner = _forced_owner()
        lead_id = kwargs.get('lead_id')
        if owner is not None and lead_id is not None:
            with get_connection() as conn:
                row = conn.execute('SELECT owner FROM leads WHERE id = %s', (lead_id,)).fetchone()
            if row is not None and (row['owner'] or '') != owner:
                return api_error('Цей лід закріплений за іншим менеджером.', 403)
        return func(*args, **kwargs)
    return wrapper

# Поля, які менеджер/адмін реально редагує з UI (решта — імпортовані дані).
_EDITABLE_FIELDS = (
    'owner', 'pipeline', 'stage', 'outreach_status', 'priority',
    'next_followup_date', 'last_touch_date', 'reply_status',
    'crm_record_id', 'sync_status', 'notes', 'manager_private_notes',
    'first_message_en',
    # Контакти й редакційний контекст не мають бути «захованими» імпортними
    # полями. Менеджер може уточнити їх руками, але джерело все одно видно в
    # картці, тож CRM не перетворюється на набір неперевірених припущень.
    'website_url', 'phone', 'whatsapp_viber', 'email', 'instagram',
    'facebook_other_social', 'source_url', 'primary_channel',
    'need_type', 'suggested_first_offer', 'why_help_fits',
)

_COLUMNS = (
    'id', 'lead_id', 'source_bucket', 'source_row_id', 'business_name', 'category',
    'country', 'city_area', 'opening_window', 'opening_date', 'status_source',
    'website_signal', 'website_url', 'primary_channel', 'phone', 'whatsapp_viber',
    'email', 'instagram', 'facebook_other_social', 'messenger_note', 'source_url',
    'priority', 'lead_score', 'contact_quality', 'owner', 'pipeline', 'stage',
    'outreach_status', 'last_touch_date', 'next_followup_date', 'followup_count',
    'reply_status', 'need_type', 'suggested_first_offer', 'why_help_fits',
    'first_message_en', 'notes', 'crm_record_id', 'sync_status', 'duplicate_key',
    'data_quality_check', 'last_file_update', 'manager_private_notes',
    'created_at', 'updated_at',
    # Результат діагностики (lead_prospector). Без цих колонок рушій писав у
    # базу, а менеджер бачив стару здогадку з website_signal: діагноз просто
    # не доходив до фронта, бо список полів старший за міграцію.
    'domain', 'domain_source', 'diagnosis', 'diagnosis_evidence',
    'has_whatsapp', 'score_why', 'checked_at',
)

# Людські назви діагнозів + чи можна з цим іти до клієнта. `sellable=False`
# означає, що факт службовий: 'blocked' — це нас не пустив фаєрвол, а не
# поломка, 'domain_unknown' — ми не знайшли адресу, а не переконались, що
# сайту немає. Перетворити їх на претензію = втратити клієнта.
DIAGNOSIS_META = {
    'dead_dns':       ('Домен не працює',        'Сайт не відкривається ні в кого',       True),
    'unreachable':    ('Сайт не відповідає',     'Сервер мовчить',                        True),
    'http_5xx':       ('Помилка сервера',        'Сайт зламаний просто зараз',            True),
    'tls_expired':    ('Сертифікат протух',      'Браузер лякає відвідувачів',            True),
    'parked':         ('Заглушка на домені',     'Домен куплено, сайту немає',            True),
    'placeholder':    ('Сторінка-заглушка',      'Сайт обіцяно, але не зроблено',         True),
    'broken_shop':    ('Магазин не працює',      'Люди не можуть купити',                 True),
    'no_shop':        ('Торгує без магазину',    'Продажів через сайт немає',             True),
    'social_only':    ('Тільки соцмережі',       'Свого сайту немає',                     True),
    'blocked':        ('Захист від ботів',       'Перевірити руками — сайт, певно, цілий', False),
    'domain_unknown': ('Адресу сайту не знайдено', 'Це НЕ означає, що сайту немає',       False),
    'ok':             ('Сайт працює',            'Технічного приводу немає',              False),
}


def insertable_cols_for_import() -> list[str]:
    """Колонки для INSERT/UPDATE при імпорті (усе, крім id/created_at/updated_at)."""
    return [c for c in _COLUMNS if c not in ('id', 'created_at', 'updated_at')]


def _now_sql() -> str:
    from ..config import USE_PG
    return 'NOW()' if USE_PG else 'CURRENT_TIMESTAMP'


def _ensure_schema() -> None:
    from ..config import USE_PG
    pk_sql = 'SERIAL PRIMARY KEY' if USE_PG else 'INTEGER PRIMARY KEY'
    now_sql = _now_sql()
    with get_connection() as conn:
        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS leads (
                id {pk_sql},
                lead_id VARCHAR(20) UNIQUE NOT NULL,
                source_bucket VARCHAR(80) NOT NULL DEFAULT '',
                source_row_id VARCHAR(20) NOT NULL DEFAULT '',
                business_name VARCHAR(200) NOT NULL,
                category VARCHAR(120) NOT NULL DEFAULT '',
                country VARCHAR(80) NOT NULL DEFAULT '',
                city_area VARCHAR(120) NOT NULL DEFAULT '',
                opening_window VARCHAR(40) NOT NULL DEFAULT '',
                opening_date VARCHAR(40),
                status_source VARCHAR(60) NOT NULL DEFAULT '',
                website_signal VARCHAR(60) NOT NULL DEFAULT '',
                website_url TEXT,
                primary_channel VARCHAR(60) NOT NULL DEFAULT '',
                phone VARCHAR(60),
                whatsapp_viber VARCHAR(60),
                email TEXT,
                instagram VARCHAR(120),
                facebook_other_social TEXT,
                messenger_note VARCHAR(120) NOT NULL DEFAULT '',
                source_url TEXT,
                priority VARCHAR(20) NOT NULL DEFAULT 'Medium',
                lead_score INTEGER NOT NULL DEFAULT 0,
                contact_quality VARCHAR(20) NOT NULL DEFAULT '',
                owner VARCHAR(80) NOT NULL DEFAULT '',
                pipeline VARCHAR(80) NOT NULL DEFAULT 'Opening leads',
                stage VARCHAR(40) NOT NULL DEFAULT 'New',
                outreach_status VARCHAR(60) NOT NULL DEFAULT 'Not contacted',
                last_touch_date VARCHAR(40),
                next_followup_date VARCHAR(40),
                followup_count INTEGER NOT NULL DEFAULT 0,
                reply_status VARCHAR(60) NOT NULL DEFAULT 'No reply yet',
                need_type VARCHAR(120) NOT NULL DEFAULT '',
                suggested_first_offer TEXT,
                why_help_fits TEXT,
                first_message_en TEXT,
                notes TEXT,
                crm_record_id VARCHAR(80),
                sync_status VARCHAR(40) NOT NULL DEFAULT 'Ready',
                duplicate_key TEXT UNIQUE,
                data_quality_check VARCHAR(40) NOT NULL DEFAULT '',
                last_file_update VARCHAR(40),
                manager_private_notes TEXT,
                created_at TIMESTAMP NOT NULL DEFAULT {now_sql},
                updated_at TIMESTAMP NOT NULL DEFAULT {now_sql}
            )
            """
        )
        conn.execute('CREATE INDEX IF NOT EXISTS idx_leads_owner ON leads(owner)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_leads_stage ON leads(stage)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_leads_pipeline ON leads(pipeline)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_leads_priority ON leads(priority)')

        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS lead_activity (
                id {pk_sql},
                lead_id INTEGER NOT NULL REFERENCES leads(id) ON DELETE CASCADE,
                author VARCHAR(80) NOT NULL DEFAULT '',
                kind VARCHAR(20) NOT NULL DEFAULT 'note',
                text TEXT NOT NULL DEFAULT '',
                created_at TIMESTAMP NOT NULL DEFAULT {now_sql}
            )
            """
        )
        conn.execute('CREATE INDEX IF NOT EXISTS idx_lead_activity_lead ON lead_activity(lead_id, created_at)')

        conn.execute(
            f"""
            CREATE TABLE IF NOT EXISTS crm_activity_log (
                id {pk_sql}, user_id INTEGER NOT NULL,
                kind VARCHAR(30) NOT NULL DEFAULT 'crm',
                title VARCHAR(180) NOT NULL, detail TEXT NOT NULL DEFAULT '',
                lead_id INTEGER, lead_name VARCHAR(200) NOT NULL DEFAULT '',
                created_at TIMESTAMP NOT NULL DEFAULT {now_sql}
            )
            """
        )
        conn.execute('CREATE INDEX IF NOT EXISTS idx_crm_activity_user ON crm_activity_log(user_id, created_at)')


def _require_admin():
    """403 якщо не адмін; повертає поточного користувача інакше."""
    user = getattr(g, 'current_user', None)
    if not user or user.get('role') not in _ADMIN_ROLES:
        return None
    return user


def _row_to_payload(row: dict[str, Any]) -> dict[str, Any]:
    payload = {col: row.get(col) for col in _COLUMNS}
    payload['intelligence'] = _lead_intelligence(payload)
    return payload


@leads_bp.get('/activity-log')
@auth_required
@role_required(*_ADMIN_ROLES)
def list_activity_log():
    _ensure_schema()
    try:
        limit = min(500, max(1, int(request.args.get('limit') or 250)))
    except (TypeError, ValueError):
        limit = 250
    with get_connection() as conn:
        rows = conn.execute(
            'SELECT id, kind, title, detail, lead_id, lead_name, created_at '
            'FROM crm_activity_log WHERE user_id = %s ORDER BY id DESC LIMIT %s',
            (g.current_user['id'], limit),
        ).fetchall()
    return jsonify({'ok': True, 'data': [
        {'server_id': r['id'], 'kind': r['kind'], 'title': r['title'], 'detail': r['detail'],
         'lead_id': r['lead_id'], 'lead_name': r['lead_name'], 'created_at': r['created_at'],
         'actor': g.current_user.get('full_name') or 'Ви'} for r in (rows or [])
    ]})


@leads_bp.post('/activity-log')
@auth_required
@role_required(*_ADMIN_ROLES)
def add_activity_log():
    _ensure_schema()
    body = request.get_json(silent=True) or {}
    title = str(body.get('title') or '').strip()[:180]
    if not title:
        return api_error('Назва дії обов’язкова.', 400)
    kind = str(body.get('kind') or 'crm').strip()[:30] or 'crm'
    detail = str(body.get('detail') or '').strip()[:1000]
    lead_name = str(body.get('lead_name') or '').strip()[:200]
    lead_id = body.get('lead_id')
    try:
        lead_id = int(lead_id) if lead_id not in (None, '', False) else None
    except (TypeError, ValueError):
        lead_id = None
    with get_connection() as conn:
        cur = conn.execute(
            'INSERT INTO crm_activity_log (user_id, kind, title, detail, lead_id, lead_name) '
            'VALUES (%s, %s, %s, %s, %s, %s)' + get_returning_id_suffix(),
            (g.current_user['id'], kind, title, detail, lead_id, lead_name),
        )
        entry_id = insert_last_id(cur)
        row = conn.execute(
            'SELECT id, kind, title, detail, lead_id, lead_name, created_at '
            'FROM crm_activity_log WHERE id = %s', (entry_id,)
        ).fetchone()
    return jsonify({'ok': True, 'data': {
        'server_id': row['id'], 'kind': row['kind'], 'title': row['title'], 'detail': row['detail'],
        'lead_id': row['lead_id'], 'lead_name': row['lead_name'], 'created_at': row['created_at'],
        'actor': g.current_user.get('full_name') or 'Ви',
    }})


@leads_bp.delete('/activity-log')
@auth_required
@role_required(*_ADMIN_ROLES)
def clear_activity_log():
    """Очищує лише журнал поточного менеджера, не зачіпаючи lead_activity."""
    _ensure_schema()
    with get_connection() as conn:
        cur = conn.execute('DELETE FROM crm_activity_log WHERE user_id = %s', (g.current_user['id'],))
        deleted = int(getattr(cur, 'rowcount', 0) or 0)
    return jsonify({'ok': True, 'data': {'deleted': deleted}})


@leads_bp.get('/activity-log/export')
@auth_required
@role_required(*_ADMIN_ROLES)
def export_activity_log():
    """Експортує всю серверну історію поточного менеджера, не лише кеш браузера."""
    _ensure_schema()
    fmt = str(request.args.get('format') or 'csv').lower()
    if fmt not in {'csv', 'json'}:
        return api_error('Підтримуються формати CSV та JSON.', 400)
    with get_connection() as conn:
        rows = conn.execute(
            'SELECT created_at, kind, title, detail, lead_name, lead_id FROM crm_activity_log '
            'WHERE user_id = %s ORDER BY id DESC', (g.current_user['id'],)
        ).fetchall()
    records = [{
        'created_at': r['created_at'], 'actor': g.current_user.get('full_name') or 'Ви',
        'kind': r['kind'], 'title': r['title'], 'lead_name': r['lead_name'],
        'lead_id': r['lead_id'], 'detail': r['detail'],
    } for r in (rows or [])]
    if fmt == 'json':
        import json
        return Response(json.dumps(records, ensure_ascii=False, default=str), mimetype='application/json',
                        headers={'Content-Disposition': 'attachment; filename="arm-crm-activity.json"'})
    output = io.StringIO()
    columns = ['created_at', 'actor', 'kind', 'title', 'lead_name', 'lead_id', 'detail']
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(records)
    return Response('\ufeff' + output.getvalue(), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': 'attachment; filename="arm-crm-activity.csv"'})


def _lead_intelligence(lead: dict[str, Any]) -> dict[str, Any]:
    """Evidence-based manager hint built only from fields stored with the lead."""
    category = str(lead.get('category') or 'Локальний бізнес').strip()
    location = ', '.join(filter(None, (
        str(lead.get('city_area') or '').strip(),
        str(lead.get('country') or '').strip(),
    )))
    opening = str(lead.get('opening_date') or lead.get('opening_window') or '').strip()
    score = max(0, min(100, int(lead.get('lead_score') or 0)))
    website_signal = str(lead.get('website_signal') or '').lower()
    source = str(lead.get('status_source') or lead.get('source_bucket') or '').strip()
    contacts = [
        label for label, value in (
            ('email', lead.get('email')),
            ('телефон', lead.get('phone') or lead.get('whatsapp_viber')),
            ('Instagram', lead.get('instagram')),
        ) if str(value or '').strip()
    ]

    description = category
    if location:
        description += f' у {location}'
    if opening:
        description += f'; у даних зазначено період відкриття {opening}'
    description += '. Підказка сформована з полів CRM і потребує короткої перевірки перед контактом.'

    reasons: list[dict[str, str]] = []
    # Діагноз іде ПЕРШИМ і витісняє слабкий website_signal: це перевірений факт
    # з доказом, а не припущення джерела.
    diagnosis = str(lead.get('diagnosis') or '').strip()
    evidence = str(lead.get('diagnosis_evidence') or '').strip()
    label, hint, sellable = DIAGNOSIS_META.get(diagnosis, ('', '', False))
    if diagnosis and label:
        body = evidence or hint
        # Доказ приходить з probe і крапкою не закінчується («…HTTP 403 — схоже,
        # захист від ботів, не поломка»), тож без цього речення злипались:
        # «…не поломка Не згадуйте це клієнту».
        if body and body[-1] not in '.!?…':
            body += '.'
        if not sellable:
            body += ' Не згадуйте це клієнту як проблему.'
        reasons.append({
            'label': 'Діагноз' if sellable else 'Діагноз (не привід)',
            'text': f'{label}. {body}',
        })
    if str(lead.get('has_whatsapp') or '') in ('1', 'True', 'true'):
        reasons.append({'label': 'WhatsApp', 'text': 'На сайті є кнопка WhatsApp — повідомлення дійде до власника.'})
    if opening:
        reasons.append({'label': 'Таймінг', 'text': f'Є сигнал відкриття: {opening}. До запуску легше запропонувати базову цифрову інфраструктуру.'})
    if diagnosis:
        pass  # перевірений діагноз вище вже сказав про сайт точніше
    elif 'no' in website_signal or 'нема' in website_signal or 'відсут' in website_signal:
        reasons.append({'label': 'Слабкий сигнал', 'text': 'У джерелі не знайдено сайт. Це варто перевірити вручну перед згадкою у повідомленні.'})
    elif lead.get('website_url'):
        reasons.append({'label': 'Сайт знайдено', 'text': 'Можна швидко переглянути сайт і почати розмову з конкретного спостереження, а не загальної пропозиції.'})
    if contacts:
        reasons.append({'label': 'Контактність', 'text': f"Доступні канали: {', '.join(contacts)}. Оберіть один основний канал і не дублюйте перше звернення."})
    else:
        reasons.append({'label': 'Потрібна перевірка', 'text': 'Прямого контакту немає: спочатку знайдіть офіційний сайт або профіль власника.'})
    if score:
        reasons.append({'label': 'Оцінка CRM', 'text': f'{score}/100 — орієнтир для сортування, а не гарантія відповіді.'})
    if source:
        reasons.append({'label': 'Походження', 'text': f'Джерело: {source}. Відкрийте першоджерело перед персоналізацією.'})

    offer = str(lead.get('suggested_first_offer') or lead.get('need_type') or '').strip()
    why = str(lead.get('why_help_fits') or '').strip()
    if not offer:
        if 'no' in website_signal or 'нема' in website_signal or 'відсут' in website_signal:
            offer = 'Швидкий сайт або сторінка запуску з контактами й аналітикою'
        elif opening:
            offer = 'Пакет цифрового запуску до дати відкриття'
        else:
            offer = 'Короткий аудит поточної цифрової присутності'
    angle = why or f'Почніть із релевантного спостереження про {category.lower()}, а потім запропонуйте: {offer}.'
    next_step = 'Перевірити джерело й один цифровий канал, знайти конкретну деталь, потім надіслати коротке персоналізоване повідомлення.'
    strength = 'high' if score >= 70 and contacts else 'medium' if score >= 40 or contacts else 'low'
    return {
        'description': description,
        'reasons': reasons[:5],
        'recommended_offer': offer,
        'outreach_angle': angle,
        'next_step': next_step,
        'strength': strength,
        'score': score,
    }


def _build_leads_filter() -> tuple[str, list]:
    """Читає owner/stage/pipeline/priority/outreach/country/channel/search
    та повертає (WHERE ..., params) — спільне для list/export."""
    owner = (request.args.get('owner') or '').strip()
    forced = _forced_owner()
    if forced is not None:
        owner = ''              # менеджер бачить своїх і нерозподілених лідів
    stage = (request.args.get('stage') or '').strip()
    pipeline = (request.args.get('pipeline') or '').strip()
    priority = (request.args.get('priority') or '').strip()
    outreach_status = (request.args.get('outreach_status') or '').strip()
    country = (request.args.get('country') or '').strip()
    channel = (request.args.get('channel') or '').strip()
    search = (request.args.get('search') or '').strip()
    due_today = request.args.get('due_today') in ('1', 'true', 'yes')

    where = []
    params: list = []
    if forced is not None:
        where.append('(owner = %s OR owner IS NULL OR owner = %s)')
        params.extend([forced, ''])
    elif owner:
        where.append('owner = %s')
        params.append(owner)
    if stage:
        where.append('stage = %s')
        params.append(stage)
    if pipeline:
        where.append('pipeline = %s')
        params.append(pipeline)
    if priority:
        where.append('priority = %s')
        params.append(priority)
    if outreach_status:
        where.append('outreach_status = %s')
        params.append(outreach_status)
    if country:
        where.append('country = %s')
        params.append(country)
    if channel:
        where.append('primary_channel = %s')
        params.append(channel)
    if due_today:
        where.append("next_followup_date IS NOT NULL AND next_followup_date != '' AND next_followup_date <= %s")
        params.append(date.today().isoformat())
    if search:
        where.append(
            '(business_name ILIKE %s OR category ILIKE %s OR city_area ILIKE %s OR country ILIKE %s)'
            if _use_pg() else
            '(business_name LIKE %s OR category LIKE %s OR city_area LIKE %s OR country LIKE %s)'
        )
        like = f'%{search}%'
        params.extend([like, like, like, like])
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
    return where_sql, params


@leads_bp.get('')
@leads_bp.get('/')
@auth_required
@role_required(*_ADMIN_ROLES)
def list_leads():
    _ensure_schema()
    page = max(1, int(request.args.get('page') or 1))
    per_page = min(200, max(1, int(request.args.get('per_page') or 50)))
    offset = (page - 1) * per_page
    where_sql, params = _build_leads_filter()
    sort = (request.args.get('sort') or 'score').strip().lower()
    order_options = {
        'score': 'lead_score DESC, id ASC',
        'followup': "CASE WHEN next_followup_date IS NULL OR next_followup_date = '' THEN 1 ELSE 0 END, next_followup_date ASC, lead_score DESC, id ASC",
        'newest': 'id DESC',
        'name': 'business_name ASC, id ASC',
    }
    if sort not in order_options:
        sort = 'score'
    order_sql = order_options[sort]

    with get_connection() as conn:
        total = (conn.execute(
            f'SELECT COUNT(*) AS n FROM leads {where_sql}', params
        ).fetchone() or {}).get('n') or 0
        rows = conn.execute(
            f"""
            SELECT * FROM leads {where_sql}
            ORDER BY {order_sql}
            LIMIT %s OFFSET %s
            """,
            params + [per_page, offset],
        ).fetchall()

    # NB: the messenger frontend's api() helper unwraps the top-level 'data'
    # field only, so pagination must be nested inside 'data' to survive that.
    return jsonify({
        'ok': True,
        'data': {
            'items': [_row_to_payload(dict(r)) for r in (rows or [])],
            'page': page, 'per_page': per_page, 'total': int(total),
            'pages': max(1, math.ceil(int(total) / per_page)), 'sort': sort,
        },
    })


def _use_pg() -> bool:
    from ..config import USE_PG
    return USE_PG


@leads_bp.get('/stats')
@auth_required
@role_required(*_ADMIN_ROLES)
def leads_stats():
    _ensure_schema()
    forced = _forced_owner()
    # для менеджера всі цифри рахуються лише по його лідах
    scope = ' WHERE owner = %s' if forced is not None else ''
    and_scope = ' AND owner = %s' if forced is not None else ''
    sp: list[Any] = [forced] if forced is not None else []
    with get_connection() as conn:
        total = (conn.execute(f'SELECT COUNT(*) AS n FROM leads{scope}', sp).fetchone() or {}).get('n') or 0
        by_owner = conn.execute(
            f'SELECT owner, COUNT(*) AS n FROM leads{scope} GROUP BY owner ORDER BY owner', sp
        ).fetchall()
        by_stage = conn.execute(
            f'SELECT stage, COUNT(*) AS n FROM leads{scope} GROUP BY stage ORDER BY n DESC', sp
        ).fetchall()
        by_priority = conn.execute(
            f'SELECT priority, COUNT(*) AS n FROM leads{scope} GROUP BY priority ORDER BY n DESC', sp
        ).fetchall()
        by_channel = conn.execute(
            f'SELECT primary_channel, COUNT(*) AS n FROM leads{scope} GROUP BY primary_channel ORDER BY n DESC', sp
        ).fetchall()
        # Країни віддаємо разом зі статистикою — селект фільтра будується з
        # реальних значень БД, а не з хардкоду (в базі 37 країн і вони змінюються).
        by_country = conn.execute(
            f'SELECT country, COUNT(*) AS n FROM leads{scope} GROUP BY country ORDER BY n DESC', sp
        ).fetchall()
        not_contacted = (conn.execute(
            f"SELECT COUNT(*) AS n FROM leads WHERE outreach_status = 'Not contacted'{and_scope}", sp
        ).fetchone() or {}).get('n') or 0
        due_today = (conn.execute(
            "SELECT COUNT(*) AS n FROM leads WHERE next_followup_date IS NOT NULL "
            f"AND next_followup_date != '' AND next_followup_date <= %s{and_scope}",
            [date.today().isoformat()] + sp,
        ).fetchone() or {}).get('n') or 0

    return jsonify({'ok': True, 'data': {
        'total': int(total),
        'not_contacted': int(not_contacted),
        'due_today': int(due_today),
        'by_owner': [{'owner': r['owner'], 'count': int(r['n'])} for r in (by_owner or [])],
        'by_stage': [{'stage': r['stage'], 'count': int(r['n'])} for r in (by_stage or [])],
        'by_priority': [{'priority': r['priority'], 'count': int(r['n'])} for r in (by_priority or [])],
        'by_channel': [{'channel': r['primary_channel'], 'count': int(r['n'])} for r in (by_channel or [])],
        'by_country': [{'country': r['country'], 'count': int(r['n'])} for r in (by_country or []) if r['country']],
    }})


@leads_bp.get('/map/us')
@auth_required
@role_required(*_ADMIN_ROLES)
def us_leads_map():
    """Read-only map data for CRM leads in the United States.

    Coordinates are matched locally from city names. No lead address or CRM
    record is sent to a third-party geocoding service.
    """
    _ensure_schema()
    forced = _forced_owner()
    where = ''
    params: list[Any] = []
    if forced is not None:
        where = 'WHERE owner = %s'
        params.append(forced)
    with get_connection() as conn:
        rows = conn.execute(
            'SELECT id, business_name, category, country, city_area, priority, '
            'lead_score, stage, outreach_status, next_followup_date '
            f'FROM leads {where} ORDER BY lead_score DESC, id ASC',
            params,
        ).fetchall()
    data = build_us_lead_map([dict(row) for row in (rows or [])])
    data['refreshed_at'] = datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')
    response = jsonify({'ok': True, 'data': data})
    # Lead priority and follow-up state must always be fresh for the map.
    response.headers['Cache-Control'] = 'private, no-store'
    return response


def _lead_date_key(value: Any) -> str:
    """Return an ISO date prefix for values stored as date or timestamp strings."""
    raw = str(value or '').strip()
    if len(raw) >= 10 and raw[4:5] == '-' and raw[7:8] == '-':
        return raw[:10]
    return ''


def _build_work_queue(rows: list[dict[str, Any]], today: str) -> dict[str, Any]:
    """Build an exclusive, manager-friendly action queue from active leads."""
    groups = {
        'overdue': [],
        'today': [],
        'hot_unscheduled': [],
        'untouched': [],
    }
    priority_rank = {'Hot': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Watch': 4}

    for row in rows:
        lead = _row_to_payload(row)
        due = _lead_date_key(lead.get('next_followup_date'))
        priority = str(lead.get('priority') or '')
        outreach = str(lead.get('outreach_status') or '')
        if due and due < today:
            key = 'overdue'
        elif due == today:
            key = 'today'
        elif not due and priority in ('Hot', 'High'):
            key = 'hot_unscheduled'
        elif not due and outreach == 'Not contacted':
            key = 'untouched'
        else:
            continue
        lead['queue_reason'] = key
        groups[key].append(lead)

    def sort_key(lead: dict[str, Any]):
        return (
            _lead_date_key(lead.get('next_followup_date')) or '9999-12-31',
            priority_rank.get(str(lead.get('priority') or ''), 9),
            -int(lead.get('lead_score') or 0),
            int(lead.get('id') or 0),
        )

    labels = {
        'overdue': ('Прострочені', 'Контакти, які вже чекали на відповідь менеджера.'),
        'today': ('На сьогодні', 'Заплановані дії, які потрібно закрити сьогодні.'),
        'hot_unscheduled': ('Гарячі без дати', 'Високий пріоритет, але наступний крок ще не запланований.'),
        'untouched': ('Ще не опрацьовані', 'Нові ліди без першого контакту та без дати.'),
    }
    result_groups = []
    summary = {}
    for key in ('overdue', 'today', 'hot_unscheduled', 'untouched'):
        items = sorted(groups[key], key=sort_key)
        summary[key] = len(items)
        label, description = labels[key]
        result_groups.append({
            'key': key,
            'label': label,
            'description': description,
            'count': len(items),
            'items': items[:50],
        })
    summary['total_actionable'] = sum(summary.values())
    return {'summary': summary, 'groups': result_groups, 'generated_on': today}


@leads_bp.get('/work-queue')
@auth_required
@role_required(*_ADMIN_ROLES)
def leads_work_queue():
    _ensure_schema()
    # "Мій день" мусить читати той самий набір фільтрів, що список і воронка
    # (owner/stage/priority/outreach_status/search) — інакше вибраний у лідах
    # статус не доїжджає сюди, і цифри розходяться між екранами.
    # Власна база лишається: закриті ліди в робочий день не потрапляють ніколи.
    filter_sql, params = _build_leads_filter()
    where = "WHERE stage NOT IN ('Won', 'Lost')"
    if filter_sql:
        where += ' AND ' + filter_sql[len('WHERE '):]
    with get_connection() as conn:
        rows = conn.execute(
            f'SELECT * FROM leads {where} ORDER BY lead_score DESC, id ASC LIMIT 1000',
            params,
        ).fetchall()
    data = _build_work_queue([dict(r) for r in (rows or [])], date.today().isoformat())
    return jsonify({'ok': True, 'data': data})


@leads_bp.get('/<int:lead_id>')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def get_lead(lead_id: int):
    _ensure_schema()
    with get_connection() as conn:
        row = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
    if not row:
        return api_error('Лід не знайдено.', 404)
    return jsonify({'ok': True, 'data': _row_to_payload(dict(row))})


_SYSTEM_TRACKED_FIELDS = {
    'stage': 'Стадия',
    'owner': 'Менеджер',
    'priority': 'Приоритет',
    'outreach_status': 'Статус контакта',
}

# Значення полів у БД лишаються англійською (сумісність з фільтрами/CSV/API),
# але текст у CRM-чаті/журналі менеджери читають російською — тому переклад
# застосовується тільки в момент формування рядка активності/повідомлення.
_VALUE_LABELS_RU = {
    'stage': {
        'New': 'Новый', 'Contacted': 'Связались', 'Replied': 'Ответил',
        'Qualified': 'Квалифицирован', 'Proposal Sent': 'Предложение отправлено',
        'Won': 'Выиграно', 'Lost': 'Проиграно',
    },
    'outreach_status': {
        'Not contacted': 'Не связывались', 'Message sent': 'Сообщение отправлено',
        'Follow-up sent': 'Напоминание отправлено', 'Call made': 'Звонок совершён',
        'No reply': 'Без ответа', 'Replied': 'Ответил',
    },
    'priority': {
        'Hot': 'Горячий', 'High': 'Высокий', 'Medium': 'Средний', 'Low': 'Низкий', 'Watch': 'Наблюдение',
    },
    'owner': {}, # populated dynamically on frontend
}


def _ru_value(field: str, value) -> str:
    text = str(value or '').strip()
    if not text:
        return '—'
    return _VALUE_LABELS_RU.get(field, {}).get(text, text)


def _log_activity(conn, lead_id: int, author: str, kind: str, text: str) -> None:
    conn.execute(
        'INSERT INTO lead_activity (lead_id, author, kind, text) VALUES (%s, %s, %s, %s)',
        (lead_id, author, kind, text),
    )


def _get_or_create_lead_conversation(conn, lead_id: int, actor_user_id: int) -> int:
    """Знаходить або створює `conversations`-рядок, привʼязаний до ліда, і
    гарантує, що actor_user_id є учасником (auto-join, ідемпотентно)."""
    from ..config import USE_PG
    row = conn.execute('SELECT id FROM conversations WHERE lead_id = %s', (lead_id,)).fetchone()
    if row:
        conv_id = int(row['id'])
    else:
        lead = conn.execute('SELECT business_name FROM leads WHERE id = %s', (lead_id,)).fetchone()
        group_name = str((lead or {}).get('business_name') or 'Лід')
        _true = True if USE_PG else 1
        cur = conn.execute(
            'INSERT INTO conversations (lead_id, is_group, group_name) VALUES (%s, %s, %s)'
            + get_returning_id_suffix(),
            (lead_id, _true, group_name),
        )
        conv_id = int(insert_last_id(cur))

    already = conn.execute(
        'SELECT id FROM conversation_participants WHERE conversation_id = %s AND user_id = %s',
        (conv_id, actor_user_id),
    ).fetchone()
    if not already:
        conn.execute(
            'INSERT INTO conversation_participants (conversation_id, user_id) VALUES (%s, %s)',
            (conv_id, actor_user_id),
        )
    return conv_id


@leads_bp.patch('/<int:lead_id>')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def update_lead(lead_id: int):
    _ensure_schema()
    _ensure_schedule_schema()
    body = request.get_json(silent=True) or {}
    updates = {k: v for k, v in body.items() if k in _EDITABLE_FIELDS}
    if not updates:
        return api_error('Немає полів для оновлення.', 400)

    author = str(g.current_user.get('full_name') or 'Адмін')
    set_sql = ', '.join(f'{k} = %s' for k in updates)
    params = list(updates.values())
    with get_connection() as conn:
        existing = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
        if not existing:
            return api_error('Лід не знайдено.', 404)
        existing = dict(existing)
        conn.execute(
            f'UPDATE leads SET {set_sql}, updated_at = {_now_sql()} WHERE id = %s',
            params + [lead_id],
        )
        # A closed lead cannot remain an actionable calendar item.
        is_closed = str(updates.get('stage') or '') in ('Won', 'Lost')
        if is_closed:
            conn.execute("DELETE FROM lead_schedule WHERE lead_id=?", (lead_id,))
            conn.execute(
                f"UPDATE leads SET next_followup_date=NULL, updated_at={_now_sql()} WHERE id=?",
                (lead_id,),
            )
        # The lead card and the planner are two views of the same next action.
        # Keep their dates in sync whenever an editor changes either the date or
        # the responsible manager; otherwise "На завтра" looks successful but
        # never appears in the shared calendar.
        if not is_closed and ('next_followup_date' in updates or 'owner' in updates):
            target_date = str(updates.get('next_followup_date', existing.get('next_followup_date')) or '').strip()
            target_owner = str(updates.get('owner', existing.get('owner')) or '').strip()
            _sync_lead_schedule(conn, lead_id, target_date, target_owner)
        changed_lines = []
        for field, label in _SYSTEM_TRACKED_FIELDS.items():
            if field in updates and str(updates[field] or '') != str(existing.get(field) or ''):
                old_val = _ru_value(field, existing.get(field))
                new_val = _ru_value(field, updates[field])
                line = f'{label}: {old_val} → {new_val}'
                _log_activity(conn, lead_id, author, 'system', line)
                changed_lines.append(line)
        if changed_lines:
            conv_id = _get_or_create_lead_conversation(conn, lead_id, g.current_user['id'])
            msg_text = '\n'.join(changed_lines)
            conn.execute(
                'INSERT INTO messages (conversation_id, sender_id, text, msg_type) VALUES (%s, %s, %s, %s)',
                (conv_id, g.current_user['id'], encrypt_message(msg_text), 'text'),
            )
            conn.execute(
                f'UPDATE conversations SET last_message_at = {_now_sql()}, last_message_text = %s WHERE id = %s',
                (msg_text[:180], conv_id),
            )
        row = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
    return jsonify({'ok': True, 'data': _row_to_payload(dict(row))})


@leads_bp.post('/<int:lead_id>/enrich')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def enrich_lead_contacts(lead_id: int):
    """Обережно доповнює *один* лід даними з його офіційного сайту.

    Ми не вигадуємо контактів і не шукаємо дані на приватних сторінках:
    сервіс читає головну та кілька очевидних contact-сторінок, поважає
    robots.txt і заповнює лише порожні поля. Такий режим дає менеджеру
    зрозумілий контроль над якістю даних, а не непрозорий масовий скрапінг.
    """
    _ensure_schema()
    with get_connection() as conn:
        row = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
    if not row:
        return api_error('Лід не знайдено.', 404)
    lead = dict(row)
    website_url = str(lead.get('website_url') or '').strip()
    if not website_url:
        return api_error('Спочатку додайте офіційний сайт — без нього перевірити контакти неможливо.', 400)

    try:
        found = website_enrichment_service.enrich_website(website_url)
    except website_enrichment_service.WebsiteEnrichmentError as exc:
        return api_error(f'Сайт не вдалося перевірити: {exc}', 502)

    # Не стираємо те, що менеджер уже вніс вручну; заповнюємо тільки порожнє.
    incoming = {
        'phone': str(found.get('phone') or '').strip(),
        'whatsapp_viber': str(found.get('whatsapp') or '').strip(),
        'email': str(found.get('email') or '').strip(),
        'instagram': str(found.get('instagram') or '').strip(),
        'facebook_other_social': str(found.get('facebook') or '').strip(),
        'website_url': str(found.get('website_url') or website_url).strip(),
    }
    updates = {key: value for key, value in incoming.items() if value and not str(lead.get(key) or '').strip()}
    if incoming['whatsapp_viber']:
        updates['has_whatsapp'] = '1'
    quality = int(found.get('contact_quality_score') or 0)
    if quality and not str(lead.get('contact_quality') or '').strip():
        updates['contact_quality'] = str(quality)
    if updates:
        updates['last_file_update'] = date.today().isoformat()
        updates['data_quality_check'] = 'site_checked'

    author = str(g.current_user.get('full_name') or 'Менеджер')
    with get_connection() as conn:
        if updates:
            set_sql = ', '.join(f'{column} = %s' for column in updates)
            conn.execute(
                f'UPDATE leads SET {set_sql}, updated_at = {_now_sql()} WHERE id = %s',
                list(updates.values()) + [lead_id],
            )
            labels = {
                'phone': 'телефон', 'whatsapp_viber': 'WhatsApp', 'email': 'email',
                'instagram': 'Instagram', 'facebook_other_social': 'Facebook',
                'website_url': 'сайт',
            }
            changed = ', '.join(labels[key] for key in updates if key in labels)
            _log_activity(conn, lead_id, author, 'system', f'Перевірено офіційний сайт; додано: {changed or "дані перевірки"}.')
        fresh = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()

    return jsonify({'ok': True, 'data': {
        'lead': _row_to_payload(dict(fresh)),
        'updated_fields': sorted(updates),
        'pages_checked': int(found.get('pages_checked') or 0),
        'cache_hit': bool(found.get('cache_hit')),
        'evidence': found.get('evidence') or [],
        'errors': found.get('errors') or [],
    }})


@leads_bp.get('/<int:lead_id>/conversation')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def get_lead_conversation(lead_id: int):
    """Повертає id реальної розмови месенджера, привʼязаної до ліда
    (створює за потреби) і приєднує поточного адміна до неї."""
    _ensure_schema()
    with get_connection() as conn:
        lead = conn.execute('SELECT id FROM leads WHERE id = %s', (lead_id,)).fetchone()
        if not lead:
            return api_error('Лід не знайдено.', 404)
        conv_id = _get_or_create_lead_conversation(conn, lead_id, g.current_user['id'])
    return jsonify({'ok': True, 'data': {'conversation_id': conv_id}})


@leads_bp.post('/<int:lead_id>/ai-draft')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def generate_ai_draft(lead_id: int):
    """Cold-outreach чернетка першого контакту на основі даних ліда —
    безкоштовна LLM-модель через OpenRouter (fallback-ланцюжок)."""
    if not openrouter_service.is_configured():
        return api_error('AI-чернетки не налаштовані (немає OPENROUTER_API_KEY).', 503)
    _ensure_schema()
    with get_connection() as conn:
        lead = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
    if not lead:
        return api_error('Лід не знайдено.', 404)

    prompt = ai_drafts.build_draft_prompt(dict(lead))
    try:
        text, model_used = openrouter_service.generate(prompt)
    except OpenRouterError as exc:
        return api_error(f'Не вдалося згенерувати чернетку: {exc.message}', 502)

    # Слабкі безкоштовні моделі інколи ігнорують частину тегованого формату
    # (напр. видають LOCAL, але забувають EN2) — НЕ ретраїмо іншою моделлю
    # заради повноти: generate() вже пробує до _MAX_ATTEMPTS моделей саме на
    # мережеві/HTTP збої, а ще один повний прохід подвоїв би найгірший час
    # відповіді за 60-секундний таймаут gunicorn-воркера (deploy/gunicorn.conf.py).
    # Часткова відповідь (лише EN або лише LOCAL) — все одно корисна, не помилка.
    data = ai_drafts.parse_draft_response(text)
    if not data['variants_en'] and not data['local']:
        return api_error('Не вдалося згенерувати чернетку: модель не повернула розпізнаваний текст.', 502)

    data['model_used'] = model_used
    return jsonify({'ok': True, 'data': data})


@leads_bp.post('/<int:lead_id>/ai-nudge')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def generate_ai_nudge(lead_id: int):
    """Follow-up-нагадування для простроченого ліда (next_followup_date у
    минулому). Той самий EN1/EN2/LOCAL формат і парсер, що й cold-outreach
    чернетка — фронтенд підставляє обраний варіант у звичайне поле вводу
    чату, менеджер сам вичитує/редагує і тисне Надіслати (нічого не йде
    клієнту автоматично)."""
    if not openrouter_service.is_configured():
        return api_error('AI-чернетки не налаштовані (немає OPENROUTER_API_KEY).', 503)
    _ensure_schema()
    with get_connection() as conn:
        lead = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
    if not lead:
        return api_error('Лід не знайдено.', 404)
    lead = dict(lead)

    days_overdue = 0
    raw_date = str(lead.get('next_followup_date') or '').strip()
    if raw_date:
        try:
            days_overdue = (date.today() - date.fromisoformat(raw_date)).days
        except ValueError:
            days_overdue = 0
    days_overdue = max(0, days_overdue)

    prompt = ai_drafts.build_nudge_prompt(lead, days_overdue)
    try:
        text, model_used = openrouter_service.generate(prompt)
    except OpenRouterError as exc:
        return api_error(f'Не вдалося згенерувати нагадування: {exc.message}', 502)

    data = ai_drafts.parse_draft_response(text)
    if not data['variants_en'] and not data['local']:
        return api_error('Не вдалося згенерувати нагадування: модель не повернула розпізнаваний текст.', 502)

    data['model_used'] = model_used
    data['days_overdue'] = days_overdue
    return jsonify({'ok': True, 'data': data})


@leads_bp.post('/<int:lead_id>/ai-analysis')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def generate_ai_analysis(lead_id: int):
    """AI-аналітика контакту для панелі 'Аналітика контакту' — замінює
    шаблонний rule-based _lead_intelligence() реальним аналізом від LLM
    для конкретного відкритого ліда (той rule-based варіант лишається як
    дешевий фолбек-прев'ю в самому списку лідів, де AI-виклик на кожен
    з ~400 записів на кожен рендер списку був би неприйнятно повільним/дорогим)."""
    if not openrouter_service.is_configured():
        return api_error('AI-аналітика не налаштована (немає OPENROUTER_API_KEY).', 503)
    _ensure_schema()
    with get_connection() as conn:
        lead = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
    if not lead:
        return api_error('Лід не знайдено.', 404)
    lead = dict(lead)

    prompt = ai_drafts.build_intelligence_prompt(lead)
    try:
        text, model_used = openrouter_service.generate(prompt, temperature=0.4, max_tokens=600)
    except OpenRouterError as exc:
        return api_error(f'Не вдалося згенерувати аналітику: {exc.message}', 502)

    data = ai_drafts.parse_intelligence_response(text)
    if not data['description'] and not data['reasons']:
        return api_error('Не вдалося згенерувати аналітику: модель не повернула розпізнаваний текст.', 502)

    data['score'] = max(0, min(100, int(lead.get('lead_score') or 0)))
    data['model_used'] = model_used
    return jsonify({'ok': True, 'data': data})


@leads_bp.post('/<int:lead_id>/ai-reply-suggestions')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def generate_ai_reply_suggestions(lead_id: int):
    """Підказки відповіді в живому чаті ліда — мовою клієнта (готово до
    відправки) + короткий англійський глос для менеджера."""
    if not openrouter_service.is_configured():
        return api_error('AI-підказки не налаштовані (немає OPENROUTER_API_KEY).', 503)
    _ensure_schema()
    body = request.get_json(silent=True) or {}
    variant_count = min(3, max(1, int(body.get('count') or 2)))

    with get_connection() as conn:
        lead = conn.execute('SELECT * FROM leads WHERE id = %s', (lead_id,)).fetchone()
        if not lead:
            return api_error('Лід не знайдено.', 404)
        conv_id = _get_or_create_lead_conversation(conn, lead_id, g.current_user['id'])
        rows = conn.execute(
            'SELECT sender_id, text, msg_type, is_deleted FROM messages '
            'WHERE conversation_id = %s ORDER BY id DESC LIMIT 10',
            (conv_id,),
        ).fetchall()

    me_id = g.current_user['id']
    history = []
    for r in reversed(rows or []):
        r = dict(r)
        if r.get('is_deleted') or r.get('msg_type') != 'text':
            continue
        text = decrypt_message(r.get('text') or '', fallback='')
        if not text:
            continue
        role = 'assistant' if int(r['sender_id']) == int(me_id) else 'user'
        history.append({'role': role, 'text': text})

    if not history:
        return api_error('У цього ліда ще немає повідомлень для контексту.', 400)

    prompt = ai_drafts.build_reply_prompt(dict(lead), history, variant_count)
    try:
        text, model_used = openrouter_service.generate(prompt)
    except OpenRouterError as exc:
        return api_error(f'Не вдалося згенерувати підказки: {exc.message}', 502)

    variants = ai_drafts.parse_reply_response(text)
    if not variants:
        return api_error('Не вдалося згенерувати підказки: модель не повернула розпізнаваний текст.', 502)

    return jsonify({'ok': True, 'data': {'variants': variants, 'model_used': model_used}})


@leads_bp.get('/<int:lead_id>/activity')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def list_lead_activity(lead_id: int):
    _ensure_schema()
    with get_connection() as conn:
        lead = conn.execute('SELECT id FROM leads WHERE id = %s', (lead_id,)).fetchone()
        if not lead:
            return api_error('Лід не знайдено.', 404)
        rows = conn.execute(
            'SELECT * FROM lead_activity WHERE lead_id = %s ORDER BY id ASC',
            (lead_id,),
        ).fetchall()
    return jsonify({'ok': True, 'data': [
        {'id': r['id'], 'author': r['author'], 'kind': r['kind'], 'text': r['text'], 'created_at': r['created_at']}
        for r in (rows or [])
    ]})


@leads_bp.post('/<int:lead_id>/activity')
@auth_required
@role_required(*_ADMIN_ROLES)
@own_lead_only
def add_lead_activity(lead_id: int):
    _ensure_schema()
    body = request.get_json(silent=True) or {}
    text = str(body.get('text') or '').strip()
    if not text:
        return api_error("Текст обов'язковий.", 400)
    author = str(g.current_user.get('full_name') or 'Адмін')
    with get_connection() as conn:
        lead = conn.execute('SELECT id FROM leads WHERE id = %s', (lead_id,)).fetchone()
        if not lead:
            return api_error('Лід не знайдено.', 404)
        _log_activity(conn, lead_id, author, 'note', text)
        conn.execute(
            f'UPDATE leads SET last_touch_date = %s, updated_at = {_now_sql()} WHERE id = %s',
            (date.today().isoformat(), lead_id),
        )
        rows = conn.execute(
            'SELECT * FROM lead_activity WHERE lead_id = %s ORDER BY id ASC',
            (lead_id,),
        ).fetchall()
    return jsonify({'ok': True, 'data': [
        {'id': r['id'], 'author': r['author'], 'kind': r['kind'], 'text': r['text'], 'created_at': r['created_at']}
        for r in (rows or [])
    ]})


@leads_bp.post('/import')
@auth_required
@role_required(*_ADMIN_ROLES)
def import_leads():
    """Bulk upsert за duplicate_key/lead_id. Використовується скриптом імпорту
    та може повторно викликатись без дублювання записів (idempotent)."""
    _ensure_schema()
    body = request.get_json(silent=True) or {}
    items = body.get('leads') or []
    if not isinstance(items, list) or not items:
        return api_error('Порожній список leads.', 400)

    insertable_cols = insertable_cols_for_import()
    created = 0
    updated = 0
    with get_connection() as conn:
        for item in items:
            lead_id = item.get('lead_id')
            if not lead_id or exclusion_reason(item):
                continue
            existing = conn.execute(
                'SELECT id FROM leads WHERE lead_id = %s', (lead_id,)
            ).fetchone()
            values = [item.get(c) for c in insertable_cols]
            if existing:
                set_sql = ', '.join(f'{c} = %s' for c in insertable_cols)
                conn.execute(
                    f'UPDATE leads SET {set_sql}, updated_at = {_now_sql()} WHERE lead_id = %s',
                    values + [lead_id],
                )
                updated += 1
            else:
                cols_sql = ', '.join(insertable_cols)
                placeholders = ', '.join(['%s'] * len(insertable_cols))
                conn.execute(
                    f'INSERT INTO leads ({cols_sql}) VALUES ({placeholders})'
                    + get_returning_id_suffix(),
                    values,
                )
                created += 1

    return jsonify({'ok': True, 'data': {'created': created, 'updated': updated}})


# ─── Inbound: the contact form on agency.munister.com.ua ──────────────
#
# The agency site is static (GitHub Pages), so the form posts here and the
# enquiry lands in the same Leads panel the team already works in, rather
# than in a mailbox nobody grooms. Public route: no auth, and therefore
# validated, rate limited and size capped like any other open door.

_INTAKE_MAX = {
    'name': 120, 'company': 160, 'email': 200, 'phone': 60,
    'service': 120, 'budget': 60, 'message': 4000, 'page': 300, 'lang': 8,
}

# A lead that says only "hi" is still a lead; a lead with no way back is not.
_INTAKE_REQUIRED = ('name', 'message')


def _clean(value: Any, limit: int) -> str:
    text = str(value or '').replace('\r\n', '\n').strip()
    # Control characters have no business in a name or an e-mail field.
    text = ''.join(ch for ch in text if ch >= ' ' or ch == '\n')
    return text[:limit]


def _looks_like_email(value: str) -> bool:
    if not value or value.count('@') != 1:
        return False
    local, _, domain = value.partition('@')
    return bool(local) and '.' in domain and ' ' not in value and not domain.startswith('.')


def _within_hours(created_at: Any, hours: int) -> bool:
    """Чи свіжий запис. Порівняння в Python, бо SQLite і Postgres віддають
    created_at по-різному (рядок проти datetime), а вікно тут не критичне."""
    from datetime import datetime, timedelta, timezone
    if created_at is None:
        return False
    stamp = created_at
    if isinstance(stamp, str):
        try:
            stamp = datetime.fromisoformat(stamp.replace('Z', '+00:00'))
        except ValueError:
            return False
    if not isinstance(stamp, datetime):
        return False
    now = datetime.now(stamp.tzinfo) if stamp.tzinfo else datetime.now()
    return (now - stamp) < timedelta(hours=hours)


@leads_bp.post('/intake')
@rate_limit(6, 3600)
def intake_lead():
    """Приймає заявку з контактної форми сайту агенції."""
    _ensure_schema()
    body = request.get_json(silent=True) or {}

    # Honeypot: a field the form keeps off-screen and a person never fills.
    # Bots are told the submission worked, because a bot that learns it was
    # caught simply comes back without the field.
    if str(body.get('website') or '').strip():
        return jsonify({'ok': True, 'data': {'received': True}})

    fields = {k: _clean(body.get(k), limit) for k, limit in _INTAKE_MAX.items()}

    for key in _INTAKE_REQUIRED:
        if not fields[key]:
            return api_error('missing_' + key, 400)

    email = fields['email']
    if email and not _looks_like_email(email):
        return api_error('bad_email', 400)
    if not email and not fields['phone']:
        return api_error('missing_contact', 400)

    now = date.today().isoformat()
    company = fields['company'] or fields['name']
    contact_line = ' · '.join(filter(None, [fields['email'], fields['phone']]))
    note_parts = [fields['message']]
    if fields['budget']:
        note_parts.append(f"Бюджет: {fields['budget']}")
    if fields['page']:
        note_parts.append(f"Сторінка: {fields['page']}")
    if fields['lang']:
        note_parts.append(f"Мова форми: {fields['lang']}")
    notes = '\n\n'.join(note_parts)

    with get_connection() as conn:
        # The same enquiry sent twice (a double click, a retried request)
        # should not become two cards. Same contact and same text within the
        # hour is treated as the one enquiry it is.
        recent = conn.execute(
            """
            SELECT id, created_at FROM leads
             WHERE source_bucket = %s AND business_name = %s AND notes = %s
             ORDER BY id DESC
            """,
            ('agency-site', company, notes),
        ).fetchone()
        if recent is not None and _within_hours(recent['created_at'], 6):
            return jsonify({'ok': True, 'data': {'received': True, 'duplicate': True}})

        lead_id = _next_lead_id(conn)
        data = {
            'lead_id': lead_id,
            'source_bucket': 'agency-site',
            'business_name': company,
            'category': fields['service'],
            'need_type': fields['service'],
            'email': fields['email'],
            'phone': fields['phone'],
            'source_url': fields['page'],
            'primary_channel': 'Email' if fields['email'] else 'Phone',
            'pipeline': 'Inbound',
            'stage': 'New',
            # Somebody who found the site and wrote is warmer than anything
            # outbound research produces, and the queue should show that.
            # The list is ordered by lead_score, so an inbound enquiry left
            # at the default 0 would sit under four hundred cold cards.
            'priority': 'Hot',
            'lead_score': 95,
            'outreach_status': 'Not contacted',
            'reply_status': 'Awaiting our reply',
            'contact_quality': 'Direct',
            'last_touch_date': now,
            'notes': notes,
        }
        cols = list(data.keys())
        cur = conn.execute(
            f"INSERT INTO leads ({', '.join(cols)}) VALUES ({', '.join(['%s'] * len(cols))})"
            + get_returning_id_suffix(),
            [data[c] for c in cols],
        )
        new_id = insert_last_id(cur)
        _log_activity(
            conn, new_id, fields['name'], 'note',
            'Заявка з форми на agency.munister.com.ua\n'
            + (f'Контакт: {contact_line}\n' if contact_line else '')
            + (f'Послуга: {fields["service"]}\n' if fields['service'] else '')
            + '\n' + fields['message'],
        )

    return jsonify({'ok': True, 'data': {'received': True, 'lead_id': lead_id}})


def _next_lead_id(conn) -> str:
    """Наступний CRM-XXXX з урахуванням найбільшого існуючого номера."""
    rows = conn.execute("SELECT lead_id FROM leads WHERE lead_id LIKE 'CRM-%'").fetchall()
    max_n = 0
    for r in (rows or []):
        raw = str(r['lead_id'] or '')
        suffix = raw.split('-', 1)[-1]
        if suffix.isdigit():
            max_n = max(max_n, int(suffix))
    return f'CRM-{max_n + 1:04d}'


_CREATE_REQUIRED = ('business_name',)
_CREATE_FIELDS = (
    'business_name', 'category', 'country', 'city_area', 'owner', 'pipeline',
    'stage', 'priority', 'outreach_status', 'phone', 'whatsapp_viber', 'email',
    'instagram', 'website_url', 'source_url', 'need_type', 'notes',
)


@leads_bp.post('')
@leads_bp.post('/')
@auth_required
@role_required(*_ADMIN_ROLES)
def create_lead():
    """Ручне створення ліда з UI (на відміну від /import — масового)."""
    _ensure_schema()
    body = request.get_json(silent=True) or {}
    for field in _CREATE_REQUIRED:
        if not str(body.get(field) or '').strip():
            return api_error(f"Поле '{field}' обов'язкове.", 400)

    data = {k: body.get(k) for k in _CREATE_FIELDS if body.get(k) not in (None, '')}
    data.setdefault('pipeline', 'Opening leads')
    data.setdefault('stage', 'New')
    data.setdefault('priority', 'Medium')
    data.setdefault('outreach_status', 'Not contacted')
    forced = _forced_owner()
    if forced is not None:
        data['owner'] = forced      # менеджер не може завести ліда на колегу

    with get_connection() as conn:
        lead_id = _next_lead_id(conn)
        data['lead_id'] = lead_id
        cols = list(data.keys())
        cols_sql = ', '.join(cols)
        placeholders = ', '.join(['%s'] * len(cols))
        cur = conn.execute(
            f'INSERT INTO leads ({cols_sql}) VALUES ({placeholders})' + get_returning_id_suffix(),
            [data[c] for c in cols],
        )
        new_id = insert_last_id(cur)
        author = str(g.current_user.get('full_name') or 'Адмін')
        _log_activity(conn, new_id, author, 'system', 'Лід створено вручну')
        row = conn.execute('SELECT * FROM leads WHERE id = %s', (new_id,)).fetchone()

    return jsonify({'ok': True, 'data': _row_to_payload(dict(row))})


@leads_bp.get('/export')
@auth_required
@role_required(*_ADMIN_ROLES)
def export_leads():
    """CSV-експорт з тими самими owner/stage/pipeline/priority/search
    фільтрами, що й список — без пагінації (усі співпадіння)."""
    _ensure_schema()
    where_sql, params = _build_leads_filter()
    with get_connection() as conn:
        rows = conn.execute(
            f'SELECT * FROM leads {where_sql} ORDER BY lead_score DESC, id ASC',
            params,
        ).fetchall()

    # This is a working export, not a database dump. Keep technical importer
    # fields (source_bucket, row IDs, internal notes) out of managers' files.
    priority_labels = {'Hot': 'Гарячий', 'High': 'Високий', 'Medium': 'Середній', 'Low': 'Низький', 'Watch': 'Спостереження'}
    stage_labels = {'New': 'Новий', 'Contacted': 'Звʼязались', 'Replied': 'Відповів', 'Qualified': 'Кваліфікований', 'Proposal Sent': 'Пропозицію надіслано', 'Won': 'Успішно', 'Lost': 'Втрачено'}
    outreach_labels = {'Not contacted': 'Не звʼязувалися', 'Message sent': 'Повідомлення надіслано', 'Follow-up sent': 'Нагадування надіслано', 'Call made': 'Дзвінок виконано', 'No reply': 'Без відповіді', 'Replied': 'Відповів'}

    def public_source(value: Any) -> str:
        source = str(value or '').strip()
        return source if source.lower().startswith(('http://', 'https://')) else ''

    def export_record(raw: dict[str, Any]) -> list[str]:
        owner = str(raw.get('owner') or '').strip()
        contacts = ' · '.join(filter(None, [
            str(raw.get('phone') or '').strip(),
            str(raw.get('whatsapp_viber') or '').strip(),
            str(raw.get('email') or '').strip(),
            str(raw.get('instagram') or '').strip(),
        ]))
        return [
            str(raw.get('lead_id') or ''),
            str(raw.get('business_name') or ''),
            ', '.join(filter(None, [str(raw.get('city_area') or '').strip(), str(raw.get('country') or '').strip()])),
            str(raw.get('category') or ''),
            contacts,
            str(raw.get('website_url') or ''),
            public_source(raw.get('source_url')),
            owner.split()[0] if owner else '',
            priority_labels.get(str(raw.get('priority') or ''), str(raw.get('priority') or '')),
            stage_labels.get(str(raw.get('stage') or ''), str(raw.get('stage') or '')),
            outreach_labels.get(str(raw.get('outreach_status') or ''), str(raw.get('outreach_status') or '')),
            str(raw.get('next_followup_date') or ''),
            str(raw.get('opening_date') or raw.get('opening_window') or ''),
            str(raw.get('lead_score') or ''),
            str(raw.get('notes') or ''),
        ]

    export_headers = [
        'ID', 'Компанія', 'Локація', 'Категорія', 'Контакти', 'Сайт', 'Джерело',
        'Відповідальний', 'Пріоритет', 'Стадія', 'Статус контакту', 'Наступна дія',
        'Відкриття', 'Оцінка', 'Нотатки',
    ]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(export_headers)
    for r in (rows or []):
        writer.writerow(export_record(dict(r)))

    return Response(
        buf.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename="leads_export.csv"'},
    )


# ── Excel: вивантаження всіх полів і завантаження назад ──────────────────────
# CSV гине на комах, переносах рядків у нотатках і кирилиці в Excel, тому
# менеджерський обмін файлами робимо в xlsx.

_XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

# Людські заголовки для колонок БД. Імпорт розуміє і їх, і технічні імена,
# тому вивантажений файл можна правити й заливати назад без перейменувань.
_EXPORT_HEADERS = {
    'lead_id': 'ID ліда', 'business_name': 'Назва бізнесу', 'category': 'Категорія',
    'country': 'Країна', 'city_area': 'Місто / район', 'opening_window': 'Вікно відкриття',
    'opening_date': 'Дата відкриття', 'status_source': 'Джерело статусу',
    'website_signal': 'Сигнал по сайту', 'website_url': 'Сайт', 'domain': 'Домен',
    'domain_source': 'Звідки домен', 'diagnosis': 'Діагноз сайту',
    'diagnosis_evidence': 'Доказ діагнозу', 'primary_channel': 'Основний канал',
    'phone': 'Телефон', 'whatsapp_viber': 'WhatsApp / Viber', 'has_whatsapp': 'Є WhatsApp',
    'email': 'Email', 'instagram': 'Instagram', 'facebook_other_social': 'Facebook / інші',
    'messenger_note': 'Нотатка по каналу', 'source_url': 'Посилання на джерело',
    'priority': 'Пріоритет', 'lead_score': 'Бали', 'score_why': 'Чому такі бали',
    'contact_quality': 'Якість контакту', 'owner': 'Відповідальний', 'pipeline': 'Пайплайн',
    'stage': 'Стадія', 'outreach_status': 'Статус контакту', 'last_touch_date': 'Останній контакт',
    'next_followup_date': 'Наступний follow-up', 'followup_count': 'К-ть follow-up',
    'reply_status': 'Відповідь', 'need_type': 'Тип потреби',
    'suggested_first_offer': 'Перша пропозиція', 'why_help_fits': 'Чому ми підходимо',
    'first_message_en': 'Перше повідомлення (EN)', 'notes': 'Нотатки',
    'manager_private_notes': 'Приватні нотатки', 'crm_record_id': 'ID у зовнішній CRM',
    'sync_status': 'Статус синхронізації', 'duplicate_key': 'Ключ дублікатів',
    'data_quality_check': 'Перевірка якості даних', 'last_file_update': 'Оновлення файлу',
    'checked_at': 'Перевірено', 'source_bucket': 'Джерело набору', 'source_row_id': 'Рядок джерела',
    'created_at': 'Створено', 'updated_at': 'Оновлено',
}

# Ширші за замовчуванням — довгі текстові поля.
_WIDE_COLS = {
    'notes', 'manager_private_notes', 'first_message_en', 'why_help_fits',
    'suggested_first_offer', 'diagnosis_evidence', 'score_why', 'website_signal',
    'source_url', 'website_url', 'duplicate_key', 'data_quality_check',
}


def _export_columns() -> list[str]:
    """Усі колонки, крім службового id. Порядок — як у _EXPORT_HEADERS,
    решта (нові поля схеми) дописуються в кінець, щоб нічого не загубити."""
    known = [c for c in _EXPORT_HEADERS if c in _COLUMNS]
    rest = [c for c in _COLUMNS if c not in known and c != 'id']
    return known + rest


# ── Робочий набір колонок: те, з чим менеджер справді працює ────────────────
# Повний експорт має 50+ колонок і годиться для правки та заливки назад, але
# для друку й швидкого перегляду потрібен короткий зріз.
_WORK_COLUMNS = (
    'lead_id', 'business_name', 'category', 'city_area', 'country',
    'phone', 'whatsapp_viber', 'email', 'instagram', 'website_url',
    'owner', 'priority', 'stage', 'outreach_status', 'next_followup_date',
    'lead_score', 'notes',
)

# Ширини колонок для PDF (в частках). Сума неважлива — нормалізується.
_PDF_WEIGHTS = {
    'lead_id': 0.7, 'business_name': 2.2, 'category': 1.3, 'city_area': 1.8,
    'country': 0.8, 'phone': 1.2, 'whatsapp_viber': 1.0, 'email': 1.6,
    'instagram': 1.3, 'website_url': 1.6, 'owner': 1.0, 'priority': 0.7,
    'stage': 0.9, 'outreach_status': 1.0, 'next_followup_date': 0.9,
    'lead_score': 0.6, 'notes': 2.4,
}


def _fetch_filtered_leads():
    """Ліди за поточними фільтрами списку — спільне для всіх експортів."""
    _ensure_schema()
    where_sql, params = _build_leads_filter()
    with get_connection() as conn:
        return conn.execute(
            f'SELECT * FROM leads {where_sql} ORDER BY lead_score DESC, id ASC',
            params,
        ).fetchall()


# Розмітка таблиці. 10pt на три десятки колонок з переносами читалось погано,
# тому базовий кегль більший, а рядки мають висоту під два рядки тексту.
_XLSX_BASE_SIZE = 11
_XLSX_ROW_HEIGHT = 34

# Колонки, де значення саме є посиланням: клікати треба по ньому, а не по
# сусідній колонці з голим URL.
_LINK_COLS = {
    'instagram': lambda v: (v if str(v).startswith('http')
                            else f'https://www.instagram.com/{str(v).lstrip("@")}/'),
    'website_url': lambda v: v if str(v).startswith('http') else f'https://{v}',
    'source_url': lambda v: v,
    'email': lambda v: f'mailto:{v}',
}


def _build_xlsx(rows, cols, title):
    """Спільний генератор xlsx для повного і робочого експорту."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title

    head_fill = PatternFill('solid', fgColor='1D4ED8')
    for i, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=_EXPORT_HEADERS.get(col, col))
        cell.font = Font(name='Arial', size=_XLSX_BASE_SIZE, bold=True, color='FFFFFF')
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = 52 if col in _WIDE_COLS else 22
    ws.row_dimensions[1].height = 38

    # Технічні імена другим рядком: саме за ними імпорт впізнає поля, навіть
    # якщо хтось перекладе або перепише людські заголовки.
    tech_fill = PatternFill('solid', fgColor='E5E7EB')
    for i, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=i, value=col)
        cell.font = Font(name='Arial', size=9, italic=True, color='6B7280')
        cell.fill = tech_fill
        cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    band = PatternFill('solid', fgColor='F1F5F9')
    link_font = Font(name='Arial', size=_XLSX_BASE_SIZE, color='1D4ED8', underline='single')
    for r_i, r in enumerate(rows or [], start=3):
        row = dict(r)
        for c_i, col in enumerate(cols, start=1):
            value = row.get(col)
            cell = ws.cell(row=r_i, column=c_i, value='' if value is None else value)
            cell.font = Font(name='Arial', size=_XLSX_BASE_SIZE)
            cell.alignment = Alignment(vertical='center', wrap_text=col in _WIDE_COLS)
            if r_i % 2 == 1:
                cell.fill = band
            maker = _LINK_COLS.get(col)
            if maker and value:
                try:
                    cell.hyperlink = maker(value)
                    cell.font = link_font
                except Exception:
                    pass       # погане значення не має ламати весь експорт
        ws.row_dimensions[r_i].height = _XLSX_ROW_HEIGHT

    ws.freeze_panes = 'C3'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
    ws.sheet_view.showGridLines = False

    # Друк: альбомно, на ширину сторінки, шапка повторюється на кожній.
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = '1:2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@leads_bp.get('/export.xlsx')
@auth_required
@role_required(*_ADMIN_ROLES)
def export_leads_xlsx():
    """Excel з тими самими фільтрами, що й список.

    scope=full (типово) — усі колонки, придатні для правки та заливки назад.
    scope=work — короткий робочий зріз.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return api_error(
            'Excel-експорт недоступний: на сервері немає openpyxl. '
            'Встановіть його (pip install openpyxl) або скористайтесь CSV.', 503)

    scope = (request.args.get('scope') or 'full').strip().lower()
    rows = _fetch_filtered_leads()
    if scope == 'work':
        cols = [c for c in _WORK_COLUMNS if c in _COLUMNS]
        name = 'leads_work.xlsx'
    else:
        cols = _export_columns()
        name = 'leads_export.xlsx'

    data = _build_xlsx(rows, cols, 'Leads')
    return Response(
        data,
        mimetype=_XLSX_MIME,
        headers={'Content-Disposition': f'attachment; filename="{name}"'},
    )


@leads_bp.get('/export.pdf')
@auth_required
@role_required(*_ADMIN_ROLES)
def export_leads_pdf():
    """Справжня таблиця в PDF замість друку сторінки браузером.

    Друк вікна давав те, що на екрані: картки, обрізані колонки й випадкові
    розриви. Тут — альбомна таблиця з повторюваною шапкою і нумерацією
    сторінок, тобто документ, який можна віддати або підшити."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            BaseDocTemplate, Frame, PageTemplate, Paragraph, Table, TableStyle,
        )
    except ImportError:
        return api_error('PDF-експорт недоступний: на сервері немає reportlab.', 503)

    from ..services.statement_service import _ensure_fonts, _f
    _ensure_fonts()

    rows = _fetch_filtered_leads()
    cols = [c for c in _WORK_COLUMNS if c in _COLUMNS]

    page_w, page_h = landscape(A4)
    margin = 10 * mm
    avail = page_w - 2 * margin

    weights = [_PDF_WEIGHTS.get(c, 1.0) for c in cols]
    total = sum(weights) or 1
    widths = [avail * w / total for w in weights]

    head_style = ParagraphStyle(
        'h', fontName=_f(True), fontSize=6.5, leading=8, textColor=colors.white)
    cell_style = ParagraphStyle(
        'c', fontName=_f(), fontSize=6.5, leading=8, textColor=colors.HexColor('#1F2937'))

    def esc(value):
        s = '' if value is None else str(value)
        return (s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))

    table_data = [[Paragraph(esc(_EXPORT_HEADERS.get(c, c)), head_style) for c in cols]]
    for r in (rows or []):
        row = dict(r)
        table_data.append([Paragraph(esc(row.get(c)), cell_style) for c in cols])

    table = Table(table_data, colWidths=widths, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D1D5DB')),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F1F5F9')))
    table.setStyle(TableStyle(style))

    generated = date.today().isoformat()
    count = len(rows or [])

    def decorate(canvas, doc):
        canvas.saveState()
        canvas.setFont(_f(True), 11)
        canvas.setFillColor(colors.HexColor('#1F2937'))
        canvas.drawString(margin, page_h - margin + 4 * mm, 'Ліди · ARM CRM')
        canvas.setFont(_f(), 7.5)
        canvas.setFillColor(colors.HexColor('#6B7280'))
        canvas.drawString(margin, page_h - margin - 1 * mm,
                          f'{generated} · записів: {count} · поточні фільтри списку')
        canvas.drawRightString(page_w - margin, 6 * mm, f'Сторінка {doc.page}')
        canvas.restoreState()

    buf = io.BytesIO()
    doc = BaseDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=margin, rightMargin=margin,
        topMargin=margin + 6 * mm, bottomMargin=margin,
        title='Ліди · ARM CRM',
    )
    frame = Frame(margin, margin, avail, page_h - 2 * margin - 6 * mm, id='body',
                  leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    doc.addPageTemplates([PageTemplate(id='p', frames=[frame], onPage=decorate)])
    doc.build([table])
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': 'attachment; filename="leads.pdf"'},
    )


def _norm_header(value: str) -> str:
    return str(value or '').strip().lower()


# Поширені підписи з чужих файлів і вигрузок. Без них заливка чужої таблиці
# падала з «не знайдено жодної відомої колонки», хоча дані там були.
_HEADER_ALIASES = {
    'business_name': ('name', 'business', 'company', 'company name', 'business name',
                      'назва', 'назва бізнесу', 'компанія', 'название', 'компания'),
    'city_area': ('city', 'town', 'city/area', 'city / metro', 'city metro', 'location',
                  'address', 'адреса', 'місто', 'город', 'локація'),
    'country': ('country name', 'країна', 'страна'),
    'category': ('type', 'industry', 'категорія', 'категория', 'ніша'),
    'phone': ('phone number', 'tel', 'telephone', 'mobile', 'телефон'),
    'whatsapp_viber': ('whatsapp', 'whats app', 'viber', 'вотсап', 'вацап'),
    'email': ('e-mail', 'mail', 'пошта', 'почта'),
    'instagram': ('ig', 'insta', 'instagram handle', 'instagram link', 'інстаграм', 'инстаграм'),
    'facebook_other_social': ('facebook', 'fb', 'social', 'соцмережі'),
    'website_url': ('website', 'site', 'url', 'web', 'сайт'),
    'owner': ('manager', 'assigned to', 'responsible', 'відповідальний', 'менеджер'),
    'priority': ('пріоритет', 'приоритет'),
    'stage': ('status', 'стадія', 'стадия', 'этап'),
    'outreach_status': ('contact status', 'статус контакту'),
    'notes': ('note', 'comment', 'comments', 'нотатки', 'заметки', 'комментарий'),
    'lead_id': ('id', 'crm id', 'lead id', 'ід', 'ид'),
    'next_followup_date': ('followup', 'follow up', 'next contact', 'наступний контакт'),
    'lead_score': ('score', 'бали', 'оценка'),
}


def _header_map() -> dict[str, str]:
    """Заголовок у файлі -> колонка БД. Приймаємо і технічне ім'я, і людський
    підпис, бо менеджер може заливати як вивантажений файл, так і свій."""
    out = {}
    for col in _COLUMNS:
        out[_norm_header(col)] = col
    for col, label in _EXPORT_HEADERS.items():
        if col in _COLUMNS:
            out[_norm_header(label)] = col
    # Аліаси додаються ОСТАННІМИ і не перетирають точні збіги, щоб
    # власний заголовок завжди вигравав у здогадки.
    for col, names in _HEADER_ALIASES.items():
        if col not in _COLUMNS:
            continue
        for name in names:
            out.setdefault(_norm_header(name), col)
    return out


def _read_upload_rows(file_storage) -> tuple[list[dict], str]:
    """Читає xlsx або csv у список словників {колонка БД: значення}."""
    name = (file_storage.filename or '').lower()
    raw = file_storage.read()
    if not raw:
        return [], 'Файл порожній.'

    hmap = _header_map()

    if name.endswith(('.xlsx', '.xlsm')):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], ('Читання Excel недоступне: на сервері немає openpyxl. '
                        'Збережіть файл як CSV.')
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            return [], f'Не вдалося прочитати Excel: {exc}'
        ws = wb.active
        table = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    else:
        try:
            text = raw.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = raw.decode('cp1251', errors='replace')
        table = list(csv.reader(io.StringIO(text)))

    if not table:
        return [], 'У файлі немає рядків.'

    # Шукаємо рядок заголовків серед перших трьох: вивантажений нами файл має
    # людський підпис у першому рядку і технічні імена в другому.
    header_idx, header, best = 0, [], -1
    for i, row in enumerate(table[:3]):
        mapped = [hmap.get(_norm_header(c)) for c in (row or [])]
        score = sum(1 for m in mapped if m)
        if score > best:
            best, header_idx, header = score, i, mapped
    if best <= 0:
        return [], ('Не знайдено жодної відомої колонки. Заголовки мають збігатися '
                    'з вивантаженим файлом (напр. "lead_id"/"ID ліда", "business_name").')

    rows = []
    for raw_row in table[header_idx + 1:]:
        if not raw_row or all(c in (None, '') for c in raw_row):
            continue
        item = {}
        for col, value in zip(header, raw_row):
            if not col or value in (None, ''):
                continue
            item[col] = str(value).strip() if not isinstance(value, (int, float)) else value
        # рядок технічних імен із нашого ж експорту — не дані
        if item.get('lead_id') == 'lead_id':
            continue
        if item:
            rows.append(item)
    return rows, ''


@leads_bp.post('/import-file')
@auth_required
@role_required(*_ADMIN_ROLES)
def import_leads_file():
    """Завантаження лідів файлом (xlsx/csv). Матчиться по lead_id: наявні
    оновлюються, нові додаються — повторна заливка того самого файлу нічого
    не дублює. Рядки без lead_id отримують наступний вільний CRM-номер."""
    _ensure_schema()
    file_storage = request.files.get('file')
    if file_storage is None:
        return api_error('Файл не передано (поле "file").', 400)

    items, err = _read_upload_rows(file_storage)
    if err:
        return api_error(err, 400)
    if not items:
        return api_error('У файлі немає рядків із даними.', 400)

    insertable = insertable_cols_for_import()
    created = updated = skipped = 0
    errors: list[str] = []

    with get_connection() as conn:
        for n, item in enumerate(items, start=1):
            name = str(item.get('business_name') or '').strip()
            lead_id = str(item.get('lead_id') or '').strip()
            if not name and not lead_id:
                skipped += 1
                continue
            reason = exclusion_reason(item)
            if reason:
                skipped += 1
                errors.append(f'рядок {n}: {reason}')
                continue

            existing = None
            if lead_id:
                existing = conn.execute(
                    'SELECT id FROM leads WHERE lead_id = %s', (lead_id,)
                ).fetchone()
            if not lead_id:
                lead_id = _next_lead_id(conn)
                item['lead_id'] = lead_id

            values = [item.get(c) for c in insertable]
            try:
                if existing:
                    set_sql = ', '.join(f'{c} = %s' for c in insertable)
                    conn.execute(
                        f'UPDATE leads SET {set_sql}, updated_at = {_now_sql()} '
                        'WHERE lead_id = %s',
                        values + [lead_id],
                    )
                    updated += 1
                else:
                    cols_sql = ', '.join(insertable)
                    placeholders = ', '.join(['%s'] * len(insertable))
                    conn.execute(
                        f'INSERT INTO leads ({cols_sql}) VALUES ({placeholders})'
                        + get_returning_id_suffix(),
                        values,
                    )
                    created += 1
            except Exception as exc:
                skipped += 1
                errors.append(f'рядок {n} ({name or lead_id}): {exc}')

    return jsonify({'ok': True, 'data': {
        'created': created, 'updated': updated, 'skipped': skipped,
        'total': len(items), 'errors': errors[:20],
    }})


# ════════════════════════════════════════════════════════════
# SALES WORKDAY SCHEDULER — 5 leads/day/manager on weekdays
# ════════════════════════════════════════════════════════════
import calendar as _calendar
from datetime import date as _date, datetime as _datetime

_SCHED_PERIODS = [
    (2026, 8), (2026, 9), (2026, 10), (2026, 11), (2026, 12),
    (2027, 1),
]
_DAILY_QUOTA = 5
_SCHEDULE_SORTS = {'priority', 'oldest', 'owner'}

def _get_active_managers(conn) -> list[str]:
    rows = conn.execute("SELECT crm_owner FROM users WHERE role = 'manager' AND crm_owner IS NOT NULL").fetchall()
    return [r['crm_owner'] for r in rows]

_PRIORITY_ORDER = {'Hot': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Watch': 4, '': 5}


def _ensure_schedule_schema() -> None:
    with get_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS lead_schedule (
                id INTEGER PRIMARY KEY,
                lead_id INTEGER NOT NULL,
                owner   VARCHAR(20) NOT NULL,
                scheduled_date DATE NOT NULL,
                slot_index INTEGER DEFAULT 1,
                status VARCHAR(20) DEFAULT 'pending',
                completed_at TIMESTAMP,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(lead_id)
            )
        """)


def _sync_lead_schedule(conn, lead_id: int, scheduled_date: str, owner: str = '') -> dict | None:
    """Synchronize a lead's visible next action with its calendar entry.

    `leads.next_followup_date` powers "Мій день" while `lead_schedule` powers
    the planner.  A lead has one next action, therefore it must have at most
    one matching schedule row as well.
    """
    target_date = str(scheduled_date or '').strip()
    existing = conn.execute(
        "SELECT id, owner, scheduled_date, slot_index FROM lead_schedule WHERE lead_id=?",
        (lead_id,),
    ).fetchone()
    if not target_date:
        if existing:
            conn.execute("DELETE FROM lead_schedule WHERE lead_id=?", (lead_id,))
        return None

    target_owner = str(owner or (existing['owner'] if existing else '') or '').strip()
    if not target_owner:
        managers = _get_active_managers(conn)
        target_owner = managers[0] if managers else ''
    if not target_owner:
        # No active manager yet: retain the lead date and let the first plan
        # generation distribute it.  We never create an ownerless schedule row.
        return None

    if existing:
        conn.execute(
            """UPDATE lead_schedule
               SET owner=?, scheduled_date=?, status='pending', completed_at=NULL, notes=''
               WHERE id=?""",
            (target_owner, target_date, existing['id']),
        )
        return {'sched_id': existing['id'], 'owner': target_owner, 'scheduled_date': target_date}

    slot_row = conn.execute(
        "SELECT COALESCE(MAX(slot_index), 0) AS slot FROM lead_schedule WHERE owner=? AND scheduled_date=?",
        (target_owner, target_date),
    ).fetchone()
    slot_index = int(slot_row['slot'] or 0) + 1
    cur = conn.execute(
        """INSERT INTO lead_schedule (lead_id, owner, scheduled_date, slot_index, status)
           VALUES (?, ?, ?, ?, 'pending')""",
        (lead_id, target_owner, target_date, slot_index),
    )
    return {'sched_id': insert_last_id(cur), 'owner': target_owner, 'scheduled_date': target_date}


def _get_august_workdays(weekdays: list[int] | None = None) -> list[str]:
    """Повертає робочі дні активного періоду за налаштуванням менеджера."""
    allowed_weekdays = set(weekdays or [1, 2, 3, 4, 5])
    result = []
    for year, month in _SCHED_PERIODS:
        days_in_month = _calendar.monthrange(year, month)[1]
        for d in range(1, days_in_month + 1):
            dt = _date(year, month, d)
            if (dt.weekday() + 1) in allowed_weekdays:   # 1=пн … 7=нд
                result.append(f'{year:04d}-{month:02d}-{d:02d}')
    return result


def _generate_for_owner(
    conn,
    owner: str,
    reset_future_only: bool = False,
    daily_quota: int = _DAILY_QUOTA,
    sort_mode: str = 'priority',
    weekdays: list[int] | None = None,
) -> dict:
    today = _date.today().isoformat()

    # Clear the mirror date before replacing pending rows.  The new rows below
    # write it again; records that no longer qualify therefore cannot stay in
    # "Мій день" with an orphaned date.
    if reset_future_only:
        conn.execute(
            """UPDATE leads SET next_followup_date=NULL, updated_at=CURRENT_TIMESTAMP
               WHERE id IN (SELECT lead_id FROM lead_schedule
                            WHERE owner=? AND status='pending' AND scheduled_date >= ?)""",
            (owner, today),
        )
        conn.execute(
            "DELETE FROM lead_schedule WHERE owner=? AND status='pending' AND scheduled_date >= ?",
            (owner, today)
        )
    else:
        conn.execute(
            """UPDATE leads SET next_followup_date=NULL, updated_at=CURRENT_TIMESTAMP
               WHERE id IN (SELECT lead_id FROM lead_schedule WHERE owner=? AND status='pending')""",
            (owner,),
        )
        conn.execute("DELETE FROM lead_schedule WHERE owner=?", (owner,))

    # Плануємо тільки нові, ще не опрацьовані ліди.
    rows = conn.execute(
        """SELECT id, priority, lead_score FROM leads
           WHERE owner=? AND stage NOT IN ('Won','Lost')
             AND outreach_status='Not contacted'
           ORDER BY id ASC""",
        (owner,)
    ).fetchall()

    # Порядок задається робочими налаштуваннями, але завжди стабільний.
    if sort_mode == 'oldest':
        leads_sorted = sorted([dict(r) for r in (rows or [])], key=lambda x: int(x.get('id') or 0))
    elif sort_mode == 'owner':
        leads_sorted = sorted([dict(r) for r in (rows or [])], key=lambda x: (x.get('owner') or '', int(x.get('id') or 0)))
    else:
        leads_sorted = sorted(
            [dict(r) for r in (rows or [])],
            key=lambda x: (_PRIORITY_ORDER.get(x.get('priority') or '', 5), -(x.get('lead_score') or 0), int(x.get('id') or 0))
        )

    # Робочі дні всього періоду.
    august_days = _get_august_workdays(weekdays)

    inserted = 0
    lead_idx = 0
    scheduled_dates = {}  # date → count

    for day_str in august_days:
        # A newly generated plan starts today; it must never create an
        # artificial backlog for dates that have already passed.
        if day_str < today:
            continue
        if lead_idx >= len(leads_sorted):
            break
        # Перевіряємо чи є вже done-записи на цей день (якщо reset_future_only)
        done_count = 0
        if reset_future_only:
            done_count_row = conn.execute(
                "SELECT COUNT(*) as c FROM lead_schedule WHERE owner=? AND scheduled_date=? AND status='done'",
                (owner, day_str)
            ).fetchone()
            done_count = done_count_row['c'] if done_count_row else 0
        quota = max(1, int(daily_quota)) - done_count
        if quota <= 0:
            continue

        day_leads = []
        slots_used = 0
        while slots_used < quota and lead_idx < len(leads_sorted):
            lead = leads_sorted[lead_idx]
            lead_idx += 1
            # Пропускаємо ліди що вже в розкладі (UNIQUE(lead_id))
            existing = conn.execute(
                "SELECT id FROM lead_schedule WHERE lead_id=?", (lead['id'],)
            ).fetchone()
            if existing:
                continue
            day_leads.append((lead['id'], slots_used + 1))
            slots_used += 1

        for (lid, slot) in day_leads:
            conn.execute(
                """INSERT OR IGNORE INTO lead_schedule
                   (lead_id, owner, scheduled_date, slot_index, status)
                   VALUES (?, ?, ?, ?, 'pending')""",
                (lid, owner, day_str, slot)
            )
            inserted += 1
            scheduled_dates[day_str] = scheduled_dates.get(day_str, 0) + 1
            conn.execute(
                f"UPDATE leads SET next_followup_date=?, updated_at={_now_sql()} WHERE id=?",
                (day_str, lid),
            )

    return {
        'owner': owner,
        'total_leads': len(leads_sorted),
        'scheduled': inserted,
        'days_covered': len(scheduled_dates),
    }


def _assign_unowned_leads_for_schedule(conn, owners: list[str]) -> int:
    """Рівномірно розподіляє нерозподілений вхідний список перед плануванням.

    Це свідомо не чіпає вже призначені, виграні або закриті ліди. Завдяки цьому
    «Сформувати план» працює й одразу після імпорту або скидання призначень.
    """
    if not owners:
        return 0
    rows = conn.execute(
        """SELECT id FROM leads
           WHERE (owner IS NULL OR TRIM(owner) = '')
             AND stage NOT IN ('Won', 'Lost')
             AND outreach_status = 'Not contacted'
           ORDER BY CASE priority
             WHEN 'Hot' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2
             WHEN 'Low' THEN 3 ELSE 4 END,
             lead_score DESC, id ASC"""
    ).fetchall()
    for index, row in enumerate(rows):
        conn.execute('UPDATE leads SET owner=?, updated_at=CURRENT_TIMESTAMP WHERE id=?', (owners[index % len(owners)], row['id']))
    return len(rows)


@leads_bp.post('/schedule/generate')
@auth_required
@role_required(*_ADMIN_ROLES)
def schedule_generate():
    """Генерує спільний план команди за налаштуваннями робочого ритму."""
    _ensure_schedule_schema()
    payload = request.get_json(silent=True) or {}
    reset_future = bool(payload.get('reset_future_only', False))
    try:
        daily_quota = min(30, max(1, int(payload.get('quota', _DAILY_QUOTA))))
    except (TypeError, ValueError):
        daily_quota = _DAILY_QUOTA
    sort_mode = str(payload.get('sort') or 'priority')
    if sort_mode not in _SCHEDULE_SORTS:
        sort_mode = 'priority'
    raw_weekdays = payload.get('weekdays')
    weekdays = sorted({int(day) for day in raw_weekdays if str(day).isdigit() and 1 <= int(day) <= 7}) if isinstance(raw_weekdays, list) else [1, 2, 3, 4, 5]
    if not weekdays:
        weekdays = [1, 2, 3, 4, 5]
    results = []
    with get_connection() as conn:
        owners = _get_active_managers(conn)
        assigned_from_inbox = _assign_unowned_leads_for_schedule(conn, owners)
        for owner in owners:
            r = _generate_for_owner(conn, owner, reset_future_only=reset_future, daily_quota=daily_quota, sort_mode=sort_mode, weekdays=weekdays)
            results.append(r)
    return jsonify({'ok': True, 'data': {'results': results, 'assigned_from_inbox': assigned_from_inbox, 'daily_quota': daily_quota, 'sort': sort_mode, 'weekdays': weekdays}})


@leads_bp.get('/schedule/progress')
@auth_required
@role_required(*_ADMIN_ROLES)
def schedule_progress():
    """Зведений прогрес за активний період; менеджер бачить лише себе."""
    _ensure_schedule_schema()
    today = _date.today().isoformat()
    with get_connection() as conn:
        owners_data = {}
        forced_owner = _forced_owner()
        owners = [forced_owner] if forced_owner is not None else _get_active_managers(conn)
        for owner in owners:
            total_row = conn.execute(
                "SELECT COUNT(*) as c FROM lead_schedule WHERE owner=?",
                (owner,),
            ).fetchone()
            total = total_row['c'] if total_row else 0
            
            done_row = conn.execute(
                "SELECT COUNT(*) as c FROM lead_schedule WHERE owner=? AND status='done'",
                (owner,)
            ).fetchone()
            done = done_row['c'] if done_row else 0
            
            today_total_row = conn.execute(
                "SELECT COUNT(*) as c FROM lead_schedule WHERE owner=? AND scheduled_date=?",
                (owner, today)
            ).fetchone()
            today_total = today_total_row['c'] if today_total_row else 0
            
            today_done_row = conn.execute(
                "SELECT COUNT(*) as c FROM lead_schedule WHERE owner=? AND scheduled_date=? AND status='done'",
                (owner, today)
            ).fetchone()
            today_done = today_done_row['c'] if today_done_row else 0
            owners_data[owner] = {
                'total': total,
                'done': done,
                'pending': total - done,
                'today_total': today_total,
                'today_done': today_done,
                'percent': round(done / total * 100, 1) if total else 0,
            }
        generated_row = conn.execute(
            "SELECT COUNT(*) as c FROM lead_schedule"
        ).fetchone()
        generated = generated_row['c'] if generated_row else 0
    return jsonify({'ok': True, 'data': {
        'generated': generated > 0,
        'owners': owners_data,
        'today': today,
        'daily_quota': _DAILY_QUOTA,
        'months': [f'{year:04d}-{month:02d}' for year, month in _SCHED_PERIODS],
    }})


@leads_bp.get('/schedule/august')
@auth_required
@role_required(*_ADMIN_ROLES)
def schedule_august():
    """Повертає календар плану: {owner -> date -> [lead, ...]}."""
    _ensure_schedule_schema()
    owner_filter = request.args.get('owner', '').strip()
    with get_connection() as conn:
        forced_owner = _forced_owner()
        q = """
            SELECT s.id as sched_id, s.lead_id, s.owner, s.scheduled_date,
                   s.slot_index, s.status, s.completed_at, s.notes,
                   l.business_name, l.priority, l.lead_score, l.stage,
                   l.outreach_status, l.phone, l.whatsapp_viber, l.email,
                   l.instagram, l.city_area, l.country, l.category,
                   l.primary_channel
            FROM lead_schedule s
            JOIN leads l ON l.id = s.lead_id
            WHERE 1=1
        """
        params = []
        if forced_owner is not None:
            q += " AND s.owner = ?"
            params.append(forced_owner)
        elif owner_filter and owner_filter in _get_active_managers(conn):
            q += " AND s.owner = ?"
            params.append(owner_filter)
        q += " ORDER BY s.scheduled_date ASC, s.slot_index ASC"
        rows = conn.execute(q, params).fetchall()
        result_owners = [forced_owner] if forced_owner is not None else _get_active_managers(conn)

    # Group by owner → date → list
    result = {}
    for owner in result_owners:
        result[owner] = {}
    for r in (rows or []):
        row = dict(r)
        owner = row['owner']
        day = row['scheduled_date']
        if owner not in result:
            result[owner] = {}
        if day not in result[owner]:
            result[owner][day] = []
        result[owner][day].append({
            'sched_id': row['sched_id'],
            'lead_id': row['lead_id'],
            'slot': row['slot_index'],
            'status': row['status'],
            'completed_at': row['completed_at'],
            'notes': row['notes'],
            'business_name': row['business_name'],
            'priority': row['priority'],
            'lead_score': row['lead_score'],
            'stage': row['stage'],
            'outreach_status': row['outreach_status'],
            'phone': row['phone'] or row['whatsapp_viber'] or '',
            'email': row['email'] or '',
            'instagram': row['instagram'] or '',
            'city_area': row['city_area'] or '',
            'country': row['country'] or '',
            'category': row['category'] or '',
            'primary_channel': row['primary_channel'] or '',
        })

    return jsonify({'ok': True, 'data': result})


@leads_bp.get('/schedule/day')
@auth_required
@role_required(*_ADMIN_ROLES)
def schedule_day():
    """Ліди конкретного дня для конкретного менеджера."""
    _ensure_schedule_schema()
    day = request.args.get('date', _date.today().isoformat()).strip()
    owner = request.args.get('owner', '').strip()
    with get_connection() as conn:
        forced_owner = _forced_owner()
        if forced_owner is not None:
            owner = forced_owner
        if not owner or owner not in _get_active_managers(conn):
            return api_error('Вкажіть чинного менеджера', 400)
        rows = conn.execute(
            """SELECT s.id as sched_id, s.lead_id, s.slot_index, s.status,
                      s.completed_at, s.notes,
                      l.business_name, l.priority, l.lead_score, l.stage,
                      l.outreach_status, l.phone, l.whatsapp_viber, l.email,
                      l.instagram, l.city_area, l.country, l.category,
                      l.primary_channel, l.website_url, l.source_url
               FROM lead_schedule s
               JOIN leads l ON l.id = s.lead_id
               WHERE s.owner=? AND s.scheduled_date=?
               ORDER BY s.slot_index ASC""",
            (owner, day)
        ).fetchall()
    items = []
    for r in (rows or []):
        row = dict(r)
        items.append({
            'sched_id': row['sched_id'],
            'lead_id': row['lead_id'],
            'slot': row['slot_index'],
            'status': row['status'],
            'completed_at': row['completed_at'],
            'notes': row['notes'],
            'business_name': row['business_name'],
            'priority': row['priority'] or 'Medium',
            'lead_score': row['lead_score'] or 0,
            'stage': row['stage'],
            'outreach_status': row['outreach_status'] or '',
            'phone': row['phone'] or row['whatsapp_viber'] or '',
            'email': row['email'] or '',
            'instagram': row['instagram'] or '',
            'city_area': row['city_area'] or '',
            'country': row['country'] or '',
            'category': row['category'] or '',
            'primary_channel': row['primary_channel'] or '',
            'website_url': row['website_url'] or '',
        })
    return jsonify({'ok': True, 'data': {'date': day, 'owner': owner, 'items': items}})


@leads_bp.patch('/schedule/<int:sched_id>/status')
@auth_required
@role_required(*_ADMIN_ROLES)
def schedule_set_status(sched_id: int):
    """Позначити лід у розкладі як done / skipped / pending."""
    _ensure_schedule_schema()
    body = request.get_json(silent=True) or {}
    status = str(body.get('status', 'done')).strip()
    notes  = str(body.get('notes', '') or '').strip()
    if status not in ('done', 'skipped', 'pending'):
        return api_error('status має бути done / skipped / pending', 400)
    completed_at = _datetime.utcnow().isoformat() if status == 'done' else None
    with get_connection() as conn:
        row = conn.execute("SELECT id, owner, lead_id, scheduled_date FROM lead_schedule WHERE id=?", (sched_id,)).fetchone()
        if not row:
            return api_error('Запис не знайдено', 404)
        forced_owner = _forced_owner()
        if forced_owner is not None and row['owner'] != forced_owner:
            return api_error('Цей пункт належить іншому менеджеру', 403)
        conn.execute(
            "UPDATE lead_schedule SET status=?, completed_at=?, notes=? WHERE id=?",
            (status, completed_at, notes, sched_id)
        )
        # Completing a task clears it from "Мій день". Returning it to pending
        # restores its date, so calendar, lead card and work queue never drift.
        if status == 'pending':
            conn.execute(
                f"UPDATE leads SET next_followup_date=?, updated_at={_now_sql()} WHERE id=?",
                (row['scheduled_date'], row['lead_id']),
            )
        elif status == 'done':
            conn.execute(
                f"UPDATE leads SET next_followup_date=NULL, last_touch_date=?, updated_at={_now_sql()} WHERE id=?",
                (_date.today().isoformat(), row['lead_id']),
            )
        else:  # skipped: it was not a contact, only remove the planned action
            conn.execute(
                f"UPDATE leads SET next_followup_date=NULL, updated_at={_now_sql()} WHERE id=?",
                (row['lead_id'],),
            )
    return jsonify({'ok': True, 'data': {'sched_id': sched_id, 'status': status}})


# ════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════
# ANALYTICS DASHBOARD — WOW Gamification 2.0
# ════════════════════════════════════════════════════════════
from datetime import datetime as _dt
@leads_bp.get('/analytics/dashboard')
@auth_required
@role_required(*_ADMIN_ROLES)
def analytics_dashboard():
    _ensure_schema()
    try:
        _ensure_schedule_schema()
    except Exception:
        pass

    stats = {}

    with get_connection() as conn:
        forced_owner = _forced_owner()
        owners = [forced_owner] if forced_owner is not None else _get_active_managers(conn)
        for owner in owners:
            total_row = conn.execute(
                "SELECT COUNT(*) as c FROM leads WHERE owner=? AND stage != 'Lost'", (owner,)
            ).fetchone()
            total = total_row['c'] if total_row else 0
            
            contacted_row = conn.execute(
                "SELECT COUNT(*) as c FROM leads WHERE owner=? AND outreach_status != 'Not contacted' AND stage != 'Lost'", (owner,)
            ).fetchone()
            contacted = contacted_row['c'] if contacted_row else 0
            
            won_row = conn.execute(
                "SELECT COUNT(*) as c FROM leads WHERE owner=? AND stage = 'Won'", (owner,)
            ).fetchone()
            won = won_row['c'] if won_row else 0
            
            activity_row = conn.execute(
                "SELECT COUNT(*) as c FROM lead_activity la JOIN leads l ON l.id = la.lead_id WHERE l.owner=?", (owner,)
            ).fetchone()
            activity_count = activity_row['c'] if activity_row else 0
            
            try:
                sched_total_row = conn.execute(
                    "SELECT COUNT(*) as c FROM lead_schedule WHERE owner=?", (owner,)
                ).fetchone()
                sched_total = sched_total_row['c'] if sched_total_row else 0
                
                sched_done_row = conn.execute(
                    "SELECT COUNT(*) as c FROM lead_schedule WHERE owner=? AND status='done'", (owner,)
                ).fetchone()
                sched_done = sched_done_row['c'] if sched_done_row else 0
            except Exception:
                sched_total, sched_done = 0, 0

            speed_rows = conn.execute(
                """SELECT la.created_at as act_time, l.created_at as lead_time 
                   FROM lead_activity la JOIN leads l ON l.id = la.lead_id
                   WHERE l.owner=? AND la.kind='system' AND la.text LIKE 'Статус контакту: Not contacted%'""",
                (owner,)
            ).fetchall()
            
            total_hours = 0
            valid_speed_leads = 0
            for r in (speed_rows or []):
                try:
                    fmt = '%Y-%m-%d %H:%M:%S'
                    act_dt = _dt.strptime(r['act_time'][:19], fmt)
                    lead_dt = _dt.strptime(r['lead_time'][:19], fmt)
                    diff_hours = (act_dt - lead_dt).total_seconds() / 3600.0
                    if diff_hours >= 0:
                        total_hours += diff_hours
                        valid_speed_leads += 1
                except Exception:
                    continue
            avg_speed_hours = round(total_hours / valid_speed_leads, 1) if valid_speed_leads > 0 else None

            contact_rate = round((contacted / total * 100), 1) if total > 0 else 0
            win_rate = round((won / total * 100), 1) if total > 0 else 0
            sched_rate = round((sched_done / sched_total * 100), 1) if sched_total > 0 else 0
            
            # POWER SCORE CALCULATION
            speed_pts = 0
            if avg_speed_hours is not None:
                speed_pts = max(0, 30 - avg_speed_hours)
            contact_pts = contact_rate * 0.3
            sched_pts = sched_rate * 0.2
            win_pts = min(20, win_rate * 2)
            
            power_score = int(speed_pts + contact_pts + sched_pts + win_pts)
            if power_score > 100: power_score = 100
            
            if power_score <= 20: rank = 'Стартова позиція'
            elif power_score <= 50: rank = 'Стабільний темп'
            elif power_score <= 75: rank = 'Сильний результат'
            else: rank = 'Лідер продажів'

            stats[owner] = {
                'total_leads': total,
                'contacted': contacted,
                'won': won,
                'contact_rate': contact_rate,
                'win_rate': win_rate,
                'activity_count': activity_count,
                'sched_total': sched_total,
                'sched_done': sched_done,
                'sched_rate': sched_rate,
                'avg_speed_hours': avg_speed_hours,
                'power_score': power_score,
                'rank': rank
            }

    best_score = max((item['power_score'] for item in stats.values()), default=0)
    speed_values = [
        item['avg_speed_hours'] for item in stats.values()
        if item['avg_speed_hours'] is not None
    ]
    best_speed = min(speed_values) if speed_values else None
    for item in stats.values():
        item['is_leader'] = best_score > 0 and item['power_score'] == best_score
        item['is_winner_speed'] = (
            best_speed is not None and item['avg_speed_hours'] == best_speed
        )
    
    today_str = _dt.now().strftime('%Y-%m-%d')
    for owner in owners:
        with get_connection() as conn:
            won_today_row = conn.execute(
                "SELECT COUNT(*) as c FROM lead_activity WHERE author=? AND text LIKE 'Стадія:%Won%' AND created_at LIKE ?",
                (owner, f"{today_str}%")
            ).fetchone()
            won_today = won_today_row['c'] if won_today_row else 0
        stats[owner]['is_on_fire'] = (won_today > 0)

    return jsonify({'ok': True, 'data': stats})
